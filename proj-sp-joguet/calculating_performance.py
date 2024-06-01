import numpy as np
from calculating_others_network import Others_Network
import heapq
import networkx as nx
import copy
import random
import optuna
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import networkx as nx
import pandas as pd
from openpyxl import load_workbook
import sys
import ast
from matplotlib.lines import Line2D


class Global:
    def __init__(self):
        # This function will define the different hyperparameters
        self.number_station = 24
        self.std_dev_walking = 4
        self.std_dev_time_taxi = 1
        self.std_dev_price_taxi = 2
        self.max_time_walking = 15
        self.walking_over_taxi = True
        self.total_number_of_people = 0
        self.steepness_walking = 4 # Factor of the sigmoid slope
        self.midpoint_walking = 1.0 # Midpoint of the sigmoid
        self.steepness_taxi = 5.81
        self.midpoint_taxi = 0.811
        self.pollution_train = 28
        self.pollution_car = 170
        self.euclidean_to_km = 12.84/388329.757
        self.price_train = 1.5
        self.waiting_time_taxi = 5
        self.cost_train_per_km = 12.5
        self.cost_per_trainline = 0 
        self.number_trials = 2
        self.number_random_train_lines = 3
        self.number_improved_train_lines =2
        self.number_frequency_train_lines = 1
        self.reduc_factor = 0.5
        self.cost_construction_per_km = 10000000

        self.data_file = 'global_company.xlsx'
        self.data = pd.DataFrame()

        with open('data/manhattan.txt', 'r') as file:
            manhattan = file.readlines()
            self.linked_station_indices = []
            self.number_linked_stations = []
            for row_index, row in enumerate(manhattan):
                row_elements = row.split(',')
                count_of_ones_in_row = sum(float(num) == 1.0 for num in row.split(','))
                self.number_linked_stations.append(count_of_ones_in_row)
                indices = [i for i, num in enumerate(row_elements) if float(num) == 1.0]
                self.linked_station_indices.append(indices)


        with open('data/SiouxFalls_net.tntp', 'r') as file:
            table_contents = file.read()
        self.lines = table_contents.strip().split('\n')

        with open('data/SiouxFalls_trips.tntp', 'r') as file:
            content = file.read()
        trips = content.strip().split('\n')
        values = []
        for line in trips:
            if ':' in line:
                values.extend([float(val.split(':')[1]) for val in line.split(';')[:-1]])
        self.values_trips = np.array(values).reshape(self.number_station, self.number_station)

        pass

    def calculate_construction_costs(self):
        # This function will calculate the construction costs of the network for the two companies
        nodes_top_company = [1, 2, 3, 4, 5, 6, 7, 8, 0]
        shared_nodes = [10, 11, 9, 15, 17]
        nodes_bottom_company = [13, 14, 12, 16, 18, 20, 21, 22, 23, 19]
        n_years = 50
        k=-1
        distance_top = 0
        distance_bottom = 0
        for sublist in self.linked_station_indices:
            k+=1
            for element in sublist:
                if element in nodes_top_company or k in nodes_top_company: 
                    distance_top += others_network.get_euclidean_distance(k, element)
                elif element in nodes_bottom_company or k in nodes_bottom_company:
                    distance_bottom += others_network.get_euclidean_distance(k, element)
                elif element in shared_nodes and k in shared_nodes:
                    distance_top += others_network.get_euclidean_distance(k, element)/2
                    distance_bottom += others_network.get_euclidean_distance(k, element)/2
                else:
                    print("Error")
                    print("element", element)
                    print("k", k)
        distance_top = distance_top * self.euclidean_to_km
        distance_bottom = distance_bottom * self.euclidean_to_km
        cost_top = distance_top * self.cost_construction_per_km/n_years/365.25
        cost_bottom = distance_bottom * self.cost_construction_per_km/n_years/365.25
        print("cost top", cost_top)
        print("cost bottom", cost_bottom)
        self.final_cost_top = cost_top
        self.final_cost_bottom = cost_bottom

    def get_costs_train_companies(self, frequency_trainlines, train_lines):
        # This function will calculate the benefits of the train companies
        k = -1
        cost_train_company = 0
        n_hours_per_day = 7 #4h of full operating (7-9 and 18-20) and 12h of 1/4 operating (6-7; 9-18 and 20-22)
        for train_line in train_lines:
            k+=1
            total_km_line = 0
            total_time_line = np.sum(train_line[2])
            for i in range(len(train_line[0])):
                init_node = train_line[0][i]
                term_node = train_line[1][i]
                euclidean = others_network.get_euclidean_distance(init_node-1, term_node-1)
                total_km_line += euclidean * self.euclidean_to_km
            ratio_km_per_min = total_km_line/total_time_line
            cost_train_company += ratio_km_per_min * self.cost_train_per_km * n_hours_per_day * 60 * 60/frequency_trainlines[k]
        cost_train_company+= len(train_lines) * self.cost_per_trainline
        return cost_train_company 

    def calculation(self, frequency_trainlines, train_lines, time_train_array, train_info_array, stations, mapping, values_trips):
        # This function will calculate the different performance indicators
        number_station = len(stations)
        self.total_number_of_people = np.sum(values_trips)
        total_waiting_time = 0
        total_price = 0
        total_pollution = 0
        total_number_people_taking_train = 0
        total_people_walking = 0
        total_people_taking_car = 0
        total_number = 0
        is_it_valid = True
        money_first_company = 0
        money_second_company = 0 
        nodes_top_company = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        shared_nodes = [10, 11, 12, 16, 18]
        nodes_bottom_company = [13, 14, 15, 17, 19, 20, 21, 22, 23, 24]
        for i in range(0, len(train_lines[0][0])):
            if train_lines[0][0][i] in nodes_top_company:
                lines_top = range(4)
                lines_bottom = range(4, 8)
                break
            if train_lines[0][0][i] in nodes_bottom_company:
                lines_top = range(4, 8)
                lines_bottom = range(4)
                break   

        for sublist in train_lines:
            sublist[3] = [0] * len(sublist[3])
        for i in stations:
            i =mapping[i]
            for j in stations:
                j= mapping[j]
                if i != j:
                    euclidean = others_network.get_euclidean_distance(i, j)
                    number_company = np.count_nonzero(train_info_array[i*number_station+j][2] == 1)
                    pourcentage, walking, time_train, price_train, time_alternative, price_alternative = self.calculate_expected_pourcentage_taking_train(i, j, time_train_array, mapping, number_company)
                    count = np.count_nonzero(train_info_array[i*number_station+j][0] != -1)
                    for k in range (count):
                        indexes_in_first_line = [index for index, num in enumerate(train_lines[train_info_array[i*number_station+j][0][k]][0]) if num == train_info_array[i*number_station+j][1][k][0]]
                        indexes_in_second_line = [index for index, num in enumerate(train_lines[train_info_array[i*number_station+j][0][k]][1]) if num == train_info_array[i*number_station+j][1][k][1]]
                        if indexes_in_second_line:
                            max_index_in_second_line = max(indexes_in_second_line)
                        if indexes_in_first_line:
                            min_index_in_first_line = min(indexes_in_first_line)
                        for l in range(min_index_in_first_line, max_index_in_second_line+1):
                            train_lines[train_info_array[i*number_station+j][0][k]][3][l] += pourcentage*values_trips[i][j]*frequency_trainlines[train_info_array[i*number_station+j][0][k]]/60
                        total_number += pourcentage*values_trips[i][j]
                    if time_train == np.inf:
                        is_it_valid = False
                    total_waiting_time += pourcentage*time_train*values_trips[i][j] + (1-pourcentage)*time_alternative*values_trips[i][j]
                    if number_company >2 or number_company == 0:
                        is_it_valid = False
                    total_price += pourcentage*price_train*values_trips[i][j] + (1-pourcentage)*price_alternative*values_trips[i][j]
                    if number_company == 1:
                        if train_info_array[i*number_station+j][2][0] == 1:
                            money_first_company += pourcentage*price_train*values_trips[i][j]
                        if train_info_array[i*number_station+j][2][1] == 1:
                            money_second_company += pourcentage*price_train*values_trips[i][j]
                    if number_company == 2:
                        total_distance_top = 0
                        total_distance_bottom = 0
                        counter_train_to_take = np.count_nonzero(train_info_array[i*number_station+j][0] != -1)
                        for k in range (counter_train_to_take):
                            indexes_in_first_line = [index for index, num in enumerate(train_lines[train_info_array[i*number_station+j][0][k]][0]) if num == train_info_array[i*number_station+j][1][k][0]]
                            indexes_in_second_line = [index for index, num in enumerate(train_lines[train_info_array[i*number_station+j][0][k]][1]) if num == train_info_array[i*number_station+j][1][k][1]]
                            if indexes_in_second_line:
                                max_index_in_second_line = max(indexes_in_second_line)
                            if indexes_in_first_line:
                                min_index_in_first_line = min(indexes_in_first_line)
                            for l in range(min_index_in_first_line, max_index_in_second_line+1):
                                if train_info_array[i*number_station+j][0][k] in lines_top:
                                    total_distance_top += others_network.get_euclidean_distance(mapping[train_lines[train_info_array[i*number_station+j][0][k]][0][l]], mapping[train_lines[train_info_array[i*number_station+j][0][k]][1][l]])
                                if train_info_array[i*number_station+j][0][k] in lines_bottom: 
                                    total_distance_bottom += others_network.get_euclidean_distance(mapping[train_lines[train_info_array[i*number_station+j][0][k]][0][l]], mapping[train_lines[train_info_array[i*number_station+j][0][k]][1][l]])
                        if is_it_valid == True:
                            ratio_full_top = (total_distance_top)/(total_distance_top + total_distance_bottom)
                            ratio_full_bottom = (total_distance_bottom)/(total_distance_top + total_distance_bottom)
                            money_first_company += pourcentage*price_train*values_trips[i][j]*ratio_full_top 
                            money_second_company += pourcentage*price_train*values_trips[i][j]*ratio_full_bottom 
                    total_pollution += pourcentage * self.pollution_train * values_trips[i][j] * euclidean * self.euclidean_to_km
                    total_number_people_taking_train += pourcentage * values_trips[i][j]
                    if walking: 
                        total_people_walking += (1-pourcentage) * values_trips[i][j]
                    if not walking:
                        total_pollution += (1 - pourcentage) * self.pollution_car * values_trips[i][j] * euclidean * self.euclidean_to_km   
                        total_people_taking_car += (1-pourcentage) * values_trips[i][j] 
        average_waiting_time = total_waiting_time/self.total_number_of_people
        average_price = total_price/self.total_number_of_people
        pourcentage_taking_train = total_number_people_taking_train/self.total_number_of_people
        pourcentage_walking = total_people_walking/self.total_number_of_people
        pourcentage_taking_car = total_people_taking_car/self.total_number_of_people

        max_people_in_train = float('-inf')  
        total_people = 0
        total_trains = 0

        for train_info in train_lines:
            num_people_in_train = np.sum(train_info[3])
            total_people += num_people_in_train
            total_trains += len(train_info[3])
            if num_people_in_train > max_people_in_train:
                max_people_in_train = num_people_in_train

        average_people_per_train = total_people / total_trains
        
        return is_it_valid, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, max_people_in_train, pourcentage_walking, pourcentage_taking_car, train_lines, money_first_company, money_second_company

    def get_array_time_train(self, frequency_trainlines, train_lines, stations, mapping, two_companies_scenario, lines_top, lines_bottom):
        # This function will calculate the time it takes to travel from node1 to node2 using a train
        number_station = len(stations)
        time_train_array = np.full((number_station, number_station), np.inf)
        train_info_array = np.full((number_station*number_station, 3, len(train_lines)*2),-1, dtype = object)
        number_trainlines = len(train_lines)
        k= -1
        for train_line in train_lines:
            k+=1
            for i in range(len(train_line[0])):
                actual_time = 0
                waiting_time = frequency_trainlines[k]/2
                for j in range(i, len(train_line[0])): 
                    init_node = train_line[0][i]
                    term_node = train_line[1][j]
                    actual_time += train_line[2][j]
                    if (actual_time + waiting_time) < time_train_array[mapping[init_node]][mapping[term_node]] and init_node != term_node:
                        time_train_array[mapping[init_node]][mapping[term_node]] = actual_time + waiting_time
                        train_info_array[(mapping[init_node])*number_station+(mapping[term_node])][0][0] = k
                        train_info_array[(mapping[init_node])*number_station+(mapping[term_node])][1][0] = (init_node, term_node)

        np.fill_diagonal(time_train_array, 0)
        for i in stations:
            distances = [np.inf] * number_station
            i = mapping[i]
            distances[i] = 0
            visited = set()
            pq = [(0, i)]
            while pq:
                current_distance, current_node = heapq.heappop(pq)
                if current_node in visited:
                    continue
                visited.add(current_node)
                for neighbor, edge_distance in enumerate(time_train_array[current_node]):
                    if edge_distance == np.inf:
                        continue
                    distance = current_distance + edge_distance
                    if distance < distances[neighbor]:
                        first_count = np.count_nonzero(train_info_array[i*number_station+current_node][0] != -1)
                        second_count = np.count_nonzero(train_info_array[current_node*number_station+neighbor][0] != -1)
                        
                        for j in range(first_count):
                            train_info_array[i*number_station+neighbor][0][j] = train_info_array[i*number_station+current_node][0][j]
                            train_info_array[i*number_station+neighbor][1][j] = train_info_array[i*number_station+current_node][1][j]
                        for j in range(second_count):
                            train_info_array[i*number_station+neighbor][0][j+first_count] = train_info_array[current_node*number_station+neighbor][0][j]
                            train_info_array[i*number_station+neighbor][1][j+first_count] = train_info_array[current_node*number_station+neighbor][1][j]
                        for j in range(first_count+second_count, number_trainlines*2):
                            train_info_array[i*number_station+neighbor][0][j] = -1
                            train_info_array[i*number_station+neighbor][1][j] = -1         
                        distances[neighbor] = distance
                        heapq.heappush(pq, (distance, neighbor))       
            for j, distance in enumerate(distances):
                time_train_array[i][j] = distance
        train_info_array[:, 2, 0] = 1
        if two_companies_scenario:
            train_info_array[:, 2, 0] = -1
            for i in range(train_info_array.shape[0]):
                element = train_info_array[i]  
                if np.isin(lines_top, element[0]).any():
                    train_info_array[i][2][0] = 1
                if np.isin(lines_bottom, element[0]).any():
                    train_info_array[i][2][1] = 1
        return time_train_array, train_info_array
    
    def calculate_expected_pourcentage_taking_train(self, node1, node2, time_train_array, mapping, number_company):
        # This function will calculate the expected pourcentage of people taking the train,
        self.walking_over_taxi = True
        ratio_due_to_price = 1
        if number_company == 2:
            ratio_due_to_price = 1
        time_train = time_train_array[node1][node2]
        time_taxi, price_taxi, time_walking = others_network.calculate_time_and_price(node1, node2)
        random_variable_walking = np.random.normal(0,self.std_dev_walking)
        random_variable_time_taxi = np.random.normal(0,self.std_dev_time_taxi)
        random_variable_price_taxi = np.random.normal(0,self.std_dev_price_taxi)  
        time_taxi += self.waiting_time_taxi
        price_walking = 0
        if time_taxi < 1: #Setting minimal times and prices to be relevant with reality
            time_taxi = 1
        if time_walking < 1:
            time_walking = 1
        if price_taxi < 3.5: 
            price_taxi = 3.5
        if time_walking>self.max_time_walking: #If the walking time is too long, the user will take a taxi
            self.walking_over_taxi = False
        if self.walking_over_taxi:
            ratio = time_train/time_walking           
            probability_walking = 1 / (1 + np.exp(-self.steepness_walking * (ratio - self.midpoint_walking)))
            return ratio_due_to_price * (1-probability_walking), True, time_train, price_taxi/4, time_walking, price_walking
        else:
            ratio = time_taxi/time_train           
            probability_train = 1 / (1 + np.exp(-self.steepness_taxi * (ratio - self.midpoint_taxi)))
            return ratio_due_to_price * probability_train, False, time_train, price_taxi/4, time_taxi, price_taxi
    
    def find_travel_time(self, x, y):
        # This function will find the travel time between two nodes
        for line in self.lines[1:]:  
            elements = line.split('\t')
            init_node = elements[1]
            term_node = elements[2]
            if (init_node == str(x) and term_node == str(y)) or (init_node == str(y) and term_node == str(x)):
                return float(elements[5])  # Free Flow Time is at index 5
        return None  # If no direct link between x and y is found 

    def generate_random_train_lines(self, number_trainlines, all_train_lines, length, nodes):
        # This function will generate the train lines randomly for the global company scenario
        train_lines = []
        valid = True
        for i in range(number_trainlines):
            starting_station = random.choice(nodes)
            length_train_line = random.randint(length[0], length[1])
            init_node = starting_station
            train_line  =[]
            for j in range(length_train_line):
                sum_passenger = 0
                next_indexes = []
                number_passenger = []
                times = []
                for k, sublist in enumerate(all_train_lines):
                    if sublist[0][0] == init_node and sublist[1][0] not in [x[0][0] for x in train_line]: 
                        if not self.is_node_in_train_lines(train_lines, sublist[1][0]):
                            next_indexes.append(sublist[1][0])
                            times.append(sublist[2][0])
                            number_passenger.append(sublist[3][0])
                if next_indexes == []:
                    for k, sublist in enumerate(all_train_lines):
                        if sublist[0][0] == init_node and sublist[1][0] not in [x[0][0] for x in train_line]:  
                            next_indexes.append(sublist[1][0])
                            times.append(sublist[2][0])
                            number_passenger.append(sublist[3][0])
                if next_indexes == []:
                    for k, sublist in enumerate(all_train_lines):
                        if sublist[0][0] == init_node:
                            next_indexes.append(sublist[1][0])
                            times.append(sublist[2][0])
                            number_passenger.append(sublist[3][0])
                    indexes_in_train_line = []
                    for idx, row in enumerate(train_line):
                        node = row[0][0]  
                        if node in next_indexes:
                            indexes_in_train_line.append(idx) 
                    min_index = min(indexes_in_train_line)
                    next_node = train_line[min_index][0][0]
                    next_node_index = next_indexes.index(next_node)
                    time = times[next_node_index]
                    row = [[init_node], [next_node], [time], [0]] 
                    train_line.append(row)
                    break
                else:
                    if len(next_indexes) > 1:
                        probabilities = number_passenger/sum(number_passenger)
                        next_node = np.random.choice(next_indexes, p = probabilities)
                        time_index = next_indexes.index(next_node)
                        time = times[time_index]
                    else:
                        next_node = next_indexes[0]
                        time = times[0]
                   
                    row = [[init_node], [next_node], [time], [0]] 
                    train_line.append(row)
                    init_node = next_node
            train_lines.append(train_line)
        train_lines = [[[subsublist[0][0] for subsublist in sublist],
                  [subsublist[1][0] for subsublist in sublist],
                  [subsublist[2][0] for subsublist in sublist]]
                 for sublist in train_lines]
        opposite_train_lines = []
        for train_line in train_lines:
            opposite_train_line = [train_line[1][::-1], train_line[0][::-1], train_line[2][::-1]]
            opposite_train_lines.append(opposite_train_line)
        train_lines.extend(opposite_train_lines)
        for sublist in train_lines:
            sublist.append([0] * len(sublist[0]))

        stations_in_train_lines = set()
        for line in train_lines:
            stations_in_train_lines.update(line[0])
        
        for i in nodes:
            if i not in stations_in_train_lines:
                valid = False
                break
        if not valid: 
            return self.generate_random_train_lines(number_trainlines, all_train_lines, length, nodes)                
                
        return train_lines

    def is_node_in_train_lines(self, train_lines_object, n):
        # This function will check if a node is in the train lines
        for train_lines in train_lines_object:
            for line in train_lines:
                if n in line[0] or n in line[1]:
                    return True
        return False

    def improve_train_lines(self, frequency_trainlines, train_lines, time_train_array, train_info_array, all_train_lines, mapping, stations, values_trips, two_companies_scenario, lines_top, lines_bottom):
        # This function will improve the train lines by trying to add and remove connections
        num_train_lines = len(train_lines)
        half_index = num_train_lines // 2
        is_it_valid,_ , _, _, pourcentage_taking_train, _, _, _, _,_, money_a,money_b = self.calculation(frequency_trainlines, train_lines, time_train_array, train_info_array, stations, mapping, values_trips)
        best_benefits = money_a + money_b - self.get_costs_train_companies(frequency_trainlines, train_lines)
        best_reversed_train_lines = copy.deepcopy(train_lines[half_index:])
        current_reversed_train_lines = copy.deepcopy(train_lines[half_index:])
        for i in range(half_index):
            improving = True
            while improving:
                improving = False
                init_node = current_reversed_train_lines[i][1][len(current_reversed_train_lines[i][0])-1]
                next_indexes = []
                times = []
                for k, sublist in enumerate(all_train_lines):
                    if sublist[0][0] == init_node and sublist[1][0] not in current_reversed_train_lines[i][0][1:] and sublist[1][0] not in current_reversed_train_lines[i][1][1:] :
                        next_indexes.append(sublist[1][0])
                        times.append(sublist[2][0])
                if next_indexes == []:
                    continue
                for j in range(len(next_indexes)):
                    row = [[init_node], [next_indexes[j]], [times[j]], [0]]
                    current_reversed_train_lines[i][0].append(init_node) 
                    current_reversed_train_lines[i][1].append(next_indexes[j])
                    current_reversed_train_lines[i][2].append(times[j])
                    current_reversed_train_lines[i][3].append(0)
                    tested_train_lines = copy.deepcopy(self.construct_reversed_train_lines(copy.deepcopy(current_reversed_train_lines)))
                    time_train_array, train_info_array = self.get_array_time_train(frequency_trainlines, tested_train_lines, stations, mapping, two_companies_scenario, lines_top, lines_bottom)
                    is_it_valid,_ , _, _, pourcentage_taking_train, _, _, _,_,_,money_a, money_b= self.calculation(frequency_trainlines, tested_train_lines, time_train_array, train_info_array, stations, mapping, values_trips)
                    current_benefits = money_a + money_b - self.get_costs_train_companies(frequency_trainlines, tested_train_lines)
                    
                    if current_benefits > best_benefits: 
                        improving = True
                        best_reversed_train_lines = copy.deepcopy(current_reversed_train_lines)
                        best_benefits = current_benefits
                        break
                    else:  
                        current_reversed_train_lines = copy.deepcopy(best_reversed_train_lines)
        network = copy.deepcopy(self.construct_reversed_train_lines(copy.deepcopy(best_reversed_train_lines)))
        best_first_half_train_lines = copy.deepcopy(network[half_index:])
        current_first_half_train_lines = copy.deepcopy(network[:half_index:])
        for i in range(half_index):
            improving = True
            while improving:
                improving = False
                init_node = current_first_half_train_lines[i][1][len(current_first_half_train_lines[i][0])-1]
                next_indexes = []
                times = []
                for k, sublist in enumerate(all_train_lines):
                    if sublist[0][0] == init_node and sublist[1][0] not in current_first_half_train_lines[i][0][1:] and sublist[1][0] not in current_first_half_train_lines[i][1][1:] :
                        next_indexes.append(sublist[1][0])
                        times.append(sublist[2][0])
                if next_indexes == []:
                    continue
                for j in range(len(next_indexes)):
                    row = [[init_node], [next_indexes[j]], [times[j]], [0]]
                    current_first_half_train_lines[i][0].append(init_node) 
                    current_first_half_train_lines[i][1].append(next_indexes[j])
                    current_first_half_train_lines[i][2].append(times[j])
                    current_first_half_train_lines[i][3].append(0)
                    a =self.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines))
                    tested_train_lines = copy.deepcopy(self.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines)))
                    time_train_array, train_info_array = self.get_array_time_train(frequency_trainlines, tested_train_lines, stations, mapping, two_companies_scenario, lines_top, lines_bottom)
                    is_it_valid,_ , _, _, pourcentage_taking_train, _, _, _,_,_ ,money_a, money_b= self.calculation(frequency_trainlines, tested_train_lines, time_train_array, train_info_array, stations, mapping, values_trips)
                    current_benefits = money_a + money_b - self.get_costs_train_companies(frequency_trainlines, tested_train_lines)
                    if current_benefits > best_benefits: 
                        improving = True
                        best_first_half_train_lines = copy.deepcopy(current_first_half_train_lines)
                        best_benefits = current_benefits
                        break
                    else:  
                        current_first_half_train_lines = copy.deepcopy(best_first_half_train_lines)
            improving = True
            while improving:
                removed_first_connection = [sublist[1:] for sublist in best_first_half_train_lines[i]]
                current_first_half_train_lines[i] = removed_first_connection
                tested_train_lines = copy.deepcopy(self.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines)))
                time_train_array, train_info_array = self.get_array_time_train(frequency_trainlines, tested_train_lines, stations, mapping,two_companies_scenario, lines_top, lines_bottom)
                is_it_valid,_ , _, _, pourcentage_taking_train, _, _, _,_,_,money_a , money_b = self.calculation(frequency_trainlines, tested_train_lines, time_train_array, train_info_array, stations, mapping, values_trips)
                current_benefits = money_a + money_b - self.get_costs_train_companies(frequency_trainlines, tested_train_lines)
                if current_benefits > best_benefits and is_it_valid:
                    best_first_half_train_lines = copy.deepcopy(current_first_half_train_lines)
                    best_benefits = current_benefits
                else:
                    current_first_half_train_lines = copy.deepcopy(best_first_half_train_lines)
                    improving = False
            improving = True
            while improving:
                removed_last_connection = [sublist[:-1] for sublist in best_first_half_train_lines[i]]
                current_first_half_train_lines[i] = removed_last_connection
                tested_train_lines = copy.deepcopy(self.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines)))
                time_train_array, train_info_array = self.get_array_time_train(frequency_trainlines, tested_train_lines, stations, mapping, two_companies_scenario, lines_top, lines_bottom)
                is_it_valid,_ , _, _, pourcentage_taking_train, _, _, _,_ ,_,money_a , money_b= self.calculation(frequency_trainlines, tested_train_lines, time_train_array, train_info_array, stations, mapping, values_trips)
                current_benefits = money_a + money_b - self.get_costs_train_companies(frequency_trainlines, tested_train_lines)
                if current_benefits > best_benefits and is_it_valid:
                    best_first_half_train_lines = copy.deepcopy(current_first_half_train_lines)
                    best_benefits = current_benefits
                else:
                    current_first_half_train_lines = copy.deepcopy(best_first_half_train_lines)
                    improving = False
 
        time_train_array, train_info_array = self.get_array_time_train(frequency_trainlines, copy.deepcopy(self.construct_reversed_train_lines(copy.deepcopy(best_first_half_train_lines))), stations, mapping, two_companies_scenario, lines_top, lines_bottom)
        return copy.deepcopy(self.construct_reversed_train_lines(copy.deepcopy(best_first_half_train_lines))), time_train_array, train_info_array

    def calculate_optimal_trajectories(self, stations, mapping):
        # This function will calculate the optimal trajectories (if all the connections are available)
        all_train_lines = []
        for line in self.lines[1:]:  
            elements = line.split('\t')
            init_node = int(elements[1])
            term_node = int(elements[2])
            row = [[init_node], [term_node], [float(elements[5])], [0]]
            all_train_lines.append(row) 
        frequency = [0]*len(all_train_lines)
        time_train_array, train_info_array= self.get_array_time_train(frequency, all_train_lines, stations, mapping, False,0,0)
        for i in range(0,self.number_station):
            for j in range(0,self.number_station):
                if i != j:
                    count = np.count_nonzero(train_info_array[i*self.number_station+j][0] != -1)
                    for k in range(count):
                        
                        indexes_in_first_line = [index for index, num in enumerate(all_train_lines[train_info_array[i*self.number_station+j][0][k]][0]) if num == train_info_array[i*self.number_station+j][1][k][0]]
                        indexes_in_second_line = [index for index, num in enumerate(all_train_lines[train_info_array[i*self.number_station+j][0][k]][1]) if num == train_info_array[i*self.number_station+j][1][k][1]]
                        if indexes_in_second_line:
                            max_index_in_second_line = max(indexes_in_second_line)
                        if indexes_in_first_line:
                            min_index_in_first_line = min(indexes_in_first_line)
                        for l in range(min_index_in_first_line, max_index_in_second_line+1):
                            all_train_lines[train_info_array[i*self.number_station+j][0][k]][3][l] += self.values_trips[i][j]
        return all_train_lines

    def construct_reversed_train_lines(self, train_lines):
        opposite_train_lines = []
        for train_line in train_lines:
            opposite_train_line = [train_line[1][::-1], train_line[0][::-1], train_line[2][::-1], train_line[3][::-1]]
            opposite_train_lines.append(opposite_train_line)
        train_lines.extend(opposite_train_lines)
        return train_lines

    def get_benefits(self, frequency_trainlines, train_lines, stations, mapping, values_trips, two_companies_scenario, lines_top, lines_bottom):
        # This function will calculate the profits of the global train company, and the revenues of the 2 companies
        time_train_array, train_info_array  = self.get_array_time_train(frequency_trainlines, train_lines, stations, mapping, two_companies_scenario, lines_top, lines_bottom)
        _,_,_,_, pourcentage_taking_train, _,_,_,_, _, money_first_company, money_second_company = self.calculation(frequency_trainlines, train_lines, time_train_array, train_info_array, stations, mapping, values_trips)
        return (money_first_company +money_second_company - self.get_costs_train_companies(frequency_trainlines, train_lines)), money_first_company, money_second_company

    def calculate_optimal_frequency(self, init_frequency, train_lines, time_train_array, train_info_array, stations, mapping, values_trips, two_companies_scenario, number_trials):
        # This function will calculate the optimal frequency of the train lines
        optuna.logging.set_verbosity(optuna.logging.ERROR)
        def objective(trial):
            frequencies = []
            for i in range(len(train_lines)//2):
                if trial.number == 0:  # First trial
                    frequency = trial.suggest_int(f'frequency_{i}', 5, 5)
                else:
                    frequency = trial.suggest_int(f'frequency_{i}', 3, 10)
                frequencies.append(frequency)
            frequencies = np.concatenate((frequencies, frequencies))
            benefits= self.get_benefits(frequencies, train_lines, stations, mapping, values_trips, two_companies_scenario, 0, 0)[0]
            return benefits
        study = optuna.create_study(direction='maximize')
        study.optimize(objective, n_trials=number_trials)
        best_frequencies = [study.best_params[f'frequency_{i}'] for i in range(len(train_lines)//2)]
        best_benefits = study.best_value
        return best_frequencies, best_benefits

    def plotting_map(self, best_train_line, best_frequency, name):
        # This function will plot the map of the train lines
        station_coordinates = {
            1: (0, 7), 2: (6, 7), 3: (0, 6), 4: (2, 6), 5: (4, 6), 6: (6, 6),
            7: (8, 5), 8: (6, 5), 9: (4, 5), 10: (4, 4), 11: (2, 4), 12: (0, 4),
            13: (0, 0), 14: (2, 2), 15: (4, 2), 16: (6, 4), 17: (6, 3), 18: (8, 4),
            19: (6, 2), 20: (6, 0), 21: (4, 0), 22: (4, 1), 23: (2, 1), 24: (2, 0)
        }
        figures = []
        i= -1
        for train_line_data in best_train_line[:len(best_train_line)]:
            i+=1
            fig = plt.figure()  # Create a new figure for each train line
            plt.title(f'Train Line {i + 1} (Frequency: {best_frequency[i]})')
            
            for station, (x, y) in station_coordinates.items():
                plt.plot(x, y, 'ko', markersize=1)  # Adjust the markersize as needed
                plt.text(x, y, station, ha='center', va='center')
            first_line = train_line_data[0]
            second_line = train_line_data[1]
            for j in range(len(first_line)):
                start_station = first_line[j]
                end_station = second_line[j]
                start_x, start_y = station_coordinates[start_station]
                end_x, end_y = station_coordinates[end_station]
                plt.plot([start_x, end_x], [start_y, end_y], color=f'C{i}')
            plt.xlim(-1, 9)
            plt.ylim(-1, 8)
            plt.xlabel('X')
            plt.ylabel('Y')
            plt.gca().set_aspect('equal', adjustable='box')
            fig.savefig(f'{name}_train_line_{i + 1}.pdf')
            figures.append(fig)  
        pass

    def calculate_the_best_train_lines(self, max_min_number_train_lines, all_train_lines, max_min_station_per_line, nodes, mapping, values_trips, number_improved_train_lines, number_frequency_train_lines, number_trials, number_random_train_lines, plotting, two_companies_scenario, lines_top, lines_bottom):
        # This function will calculate the best network for a given scenario
        random_train_lines = [self.generate_random_train_lines(max_min_number_train_lines, all_train_lines, max_min_station_per_line, nodes) for _ in range(number_random_train_lines)]
        frequency_trainlines = [5] * (max_min_number_train_lines * 2)
        time_train_arrays = []
        train_info_arrays = []
        for train_lines in random_train_lines:
            time_train_array, train_info_array = self.get_array_time_train(frequency_trainlines, train_lines, nodes, mapping, two_companies_scenario, lines_top, lines_bottom)
            time_train_arrays.append(time_train_array)
            train_info_arrays.append(train_info_array)
        benefits = [self.get_benefits(frequency_trainlines, train_lines, nodes, mapping, values_trips, two_companies_scenario, lines_top, lines_bottom)[0] for train_lines, time_train_array, train_info_array in zip(random_train_lines, time_train_arrays, train_info_arrays)]

        if plotting:
            plt.plot(benefits, label='Benefits Random train lines')
            plt.plot(sorted(benefits, reverse = True), label='Sorted Benefits')
            plt.xlabel('Index')
            plt.ylabel('Benefits')
            plt.title('Plot of Benefits from Random Train Lines')
            plt.legend()
            plt.grid(True)
            plt.savefig('benefits_random_train_lines.pdf')
        print()
        print("Calculating among the nodes: ", nodes)
        print()
        print("Current maximum benefits:", max(benefits))
        print()
        top_indices = sorted(range(len(benefits)), key=lambda i: benefits[i], reverse=True)[:number_improved_train_lines]
        best_random_train_lines = [random_train_lines[i] for i in top_indices]
        best_time_train_arrays = [time_train_arrays[i] for i in top_indices]
        best_train_info_arrays = [train_info_arrays[i] for i in top_indices]

        print("Improving the best train lines...")
        print()
        improved_train_lines = []
        improved_time_train_arrays = []
        improved_train_info_arrays = []
        for train_lines, time_train_array, train_info_array in zip(best_random_train_lines, best_time_train_arrays, best_train_info_arrays):
            improved_train_line, time_train_array, train_info_array = self.improve_train_lines(frequency_trainlines, train_lines, time_train_array, train_info_array, all_train_lines, mapping, nodes, values_trips, two_companies_scenario, lines_top, lines_bottom)
            improved_train_lines.append(improved_train_line)
            improved_time_train_arrays.append(time_train_array)
            improved_train_info_arrays.append(train_info_array)
             
        benefits = [self.get_benefits(frequency_trainlines, train_lines, nodes, mapping, values_trips, two_companies_scenario, lines_top, lines_bottom)[0] for train_lines, time_train_array, train_info_array in zip(improved_train_lines, improved_time_train_arrays, improved_train_info_arrays)]

        top_indices = sorted(range(len(improved_train_lines)), key=lambda i: benefits[i], reverse=True)[:number_frequency_train_lines]
        best_improved_train_lines = [improved_train_lines[i] for i in top_indices]
        best_time_train_arrays = [improved_time_train_arrays[i] for i in top_indices]
        best_train_info_arrays = [improved_train_info_arrays[i] for i in top_indices]

        print("Maximum benefits train company:", max(benefits))
        print()
        if plotting:
            plt.figure()
            plt.plot(benefits, label='Benefits Improved train lines')
            plt.plot(sorted(benefits, reverse = True), label='Sorted Benefits')
            plt.xlabel('Index')
            plt.ylabel('Benefits')
            plt.title('Plot of Benefits from Improved Train Lines')
            plt.legend()
            plt.grid(True)
            plt.savefig('benefits_improved_train_lines.pdf')

        print("Optimizing the frequencies...")
        print()
        optimal_frequencies = []
        best_benefits = []
        for train_lines, time_train_array, train_info_array in zip(best_improved_train_lines, best_time_train_arrays, best_train_info_arrays):
            frequencies, benefits = self.calculate_optimal_frequency(frequency_trainlines, train_lines, time_train_array, train_info_array, nodes, mapping, values_trips,two_companies_scenario, number_trials)
            optimal_frequencies.append(frequencies)
            best_benefits.append(benefits)
        print("Maximum benefits train company:", max(best_benefits))
        print()

        if plotting:
            plt.figure()
            plt.plot(best_benefits, label='Benefits Optimal Frequencies train lines')
            plt.plot(sorted(best_benefits, reverse = True), label='Sorted Benefits')
            plt.xlabel('Index')
            plt.ylabel('Benefits')
            plt.title('Plot of Benefits from Optimal Frequencies Train Lines')
            plt.legend()
            plt.grid(True)
            plt.savefig('benefits_optimal_frequencies_train_lines.pdf')

        best_index = best_benefits.index(max(best_benefits))
        best_train_line = best_improved_train_lines[best_index]
        best_time_train_array = best_time_train_arrays[best_index]
        best_train_info_array = best_train_info_arrays[best_index]
        best_frequency = optimal_frequencies[best_index]
        best_time_train_array, best_train_info_array = self.get_array_time_train(best_frequency + best_frequency, best_train_line, nodes, mapping, two_companies_scenario, lines_top, lines_bottom)
        return best_train_line, best_frequency, best_time_train_array, best_train_info_array, max(best_benefits)

    def dijkstra_shortest_path(self, graph, start, end):
        # This function will find the shortest path between two nodes
        path = nx.dijkstra_path(graph, start, end, weight='time')
        total_time = nx.dijkstra_path_length(graph, start, end, weight='time')
        total_passengers = sum(graph[u][v]['passengers'] for u, v in zip(path[:-1], path[1:]))
        return path, total_time, total_passengers
    
    def find_transition_nodes(self, optimal_path, nodes_top_company, shared_nodes, nodes_bottom_company):
        # This function will find the transition nodes between the two companies
        top_to_shared = None
        shared_to_bottom = None
        current_group = None
        first_shared_node = None
        last_shared_node = None
        counter = 0
        for i in range(len(optimal_path) - 1):
            current_node = optimal_path[i]
            next_node = optimal_path[i + 1]
            if (current_node in nodes_top_company and next_node in shared_nodes) or (current_node in nodes_bottom_company and next_node in shared_nodes):
                first_shared_node = next_node
                counter +=1
            if (current_node in shared_nodes and next_node in nodes_bottom_company) or (current_node in shared_nodes and next_node in nodes_top_company):
                last_shared_node = current_node
                counter +=1
        if counter >2:
            raise ValueError("There are more than one shared nodes in the optimal path")
        return first_shared_node, last_shared_node
        
    def global_to_excel(self, data_file, benefits_train_company, profits, costs, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, pourcentage_walking, pourcentage_taking_car, iterations, train_lines, best_frequency):
        # This function will write the results of the global company to an excel file
        new_data = pd.DataFrame({
            "Iterations": [iterations],
            "Benefits_train_company [$]": [profits],
            "Profits" : [profits - costs - self.final_cost_top - self.final_cost_bottom],
            "Costs" : [costs + self.final_cost_top + self.final_cost_bottom],
            "Percentage_taking_train": [pourcentage_taking_train * 100],
            "Average_waiting_time [min]": [average_waiting_time],
            "Average_price [$]": [average_price],
            "Total_pollution [T]": [total_pollution / 1000000],
            "Percentage_walking": [pourcentage_walking * 100],
            "Percentage_taking_car": [pourcentage_taking_car * 100],
            "Frequency": [best_frequency],
            "Train lines" : [train_lines]
                    })
        self.data = pd.concat([self.data, new_data], ignore_index=True)
        self.data.sort_values(by="Benefits_train_company [$]", ascending=False, inplace=True)

        self.data.to_excel(data_file, index=False)
        self.set_column_widths(data_file)

    def iteration_to_excel(self, data_file, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, pourcentage_walking, pourcentage_taking_car, benefits_global, top_frequency, bottom_frequency, top_lines, bottom_lines, ite, top_benef, top_costs, bottom_benef, bottom_costs):
        # This function will write the results of the 2 company scenarios to an excel file
        new_data = pd.DataFrame({
            "Final iteration": [ite],
            "Profits_top_company": [top_benef - top_costs], 
            "Benefits_first_company": [top_benef],
            "Costs_top_company": [top_costs ],
            "Profits_bottom_company": [bottom_benef - bottom_costs], 
            "Benefits_second_company": [bottom_benef],
            "Costs_bottom_company": [bottom_costs ],
            "Total_Benefits [$]": [top_benef + bottom_benef - top_costs - bottom_costs],
            "Percentage_taking_train": [pourcentage_taking_train * 100],
            "Average_waiting_time [min]": [average_waiting_time],
            "Average_price [$]": [average_price],
            "Total_pollution [T]": [total_pollution / 1000000],
            "Percentage_walking": [pourcentage_walking * 100],
            "Percentage_taking_car": [pourcentage_taking_car * 100],
            "Top Company frequency": [top_frequency],
            "Bottom Company frequency": [bottom_frequency],
            "Top Company Lines": [top_lines],
            "Bottom Company Lines" : [bottom_lines]
                    })
                    
        self.data = pd.concat([self.data, new_data], ignore_index=True)
        self.data.to_excel(data_file, index=False)
        self.set_column_widths(data_file)

    def clear_excel(self, filename):
        # This function will clear the excel file
        wb = load_workbook(filename)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        wb.save(filename)

    def load_existing_data(self, data_file):
        # This function will load the existing data from an excel file
        try:
            self.data = pd.read_excel(data_file)
        except FileNotFoundError:
            self.data = pd.DataFrame()

    def set_column_widths(self, data_file):
        # This function will set the column widths of the excel file
        wb = load_workbook(data_file)
        ws = wb.active
        column_widths = {
            "A": 25,  
            "B": 25,
            "C": 25,
            "D": 25,
            "E": 25,
            "F": 25,
            "G": 25,
            "H": 25,
            "I": 25,
            "J": 25,
            "K": 25,
            "L": 25,
            "M": 25,
            "N": 25
        }
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width
        wb.save(data_file)
    
    def plotting_graph(self):
        # This function will plot the graphs
        labels = [ 'Two_companies', 'Cooperation', 'One_company']
        values = [pd.read_excel('Two_companies_iteration.xlsx')['Total_Benefits [$]'].tolist()[0], pd.read_excel('Cooperation_scenario.xlsx')['Total_Benefits [$]'].tolist()[0], pd.read_excel('Global_company.xlsx')['Profits'].tolist()[0]]
        plt.figure(figsize=(11, 6))
        plt.bar(labels, values, color=['red', 'blue', 'green'])
        plt.title('Benefits of the Train Companies')
        plt.ylabel('Benefits in dollar')
        plt.savefig('Benefits.pdf')
        total_benefits_two_companies = pd.read_excel('Two_companies_iteration.xlsx')['Profits_top_company'].tolist()[0]
        additional_benefits_two_companies = pd.read_excel('Two_companies_iteration.xlsx')['Profits_bottom_company'].tolist()[0]
        total_benefits_cooperation = pd.read_excel('Cooperation_scenario.xlsx')['Profits_top_company'].tolist()[0]
        additional_benefits_cooperation = pd.read_excel('Cooperation_scenario.xlsx')['Profits_bottom_company'].tolist()[0]
        total_benefits_one_company = pd.read_excel('Global_company.xlsx')['Profits'].tolist()[0]

        labels = ['Two_companies', 'Cooperation', 'One_company']
        total_values = [total_benefits_two_companies, total_benefits_cooperation, total_benefits_one_company]
        additional_values = [additional_benefits_two_companies, additional_benefits_cooperation, 0]  # No additional values for the third scenario
        plt.figure(figsize=(11, 6))
        bar_width = 0.35
        p1 = plt.bar(labels[:2], total_values[:2], bar_width, color='red', label='Benefits first company')
        p2 = plt.bar(labels[:2], additional_values[:2], bar_width, bottom=total_values[:2], color='blue', label='Benefits second companys')
        p3 = plt.bar(labels[2], total_values[2], bar_width, color='green', label='Benefits (One Company)')
        plt.title('Benefits of the Train Companies')
        plt.ylabel('Benefits in dollars')
        plt.legend()
        plt.savefig('Stacked_benefits.pdf')

        labels = [ 'Two_companies', 'Cooperation', 'One_company']
        values = [pd.read_excel('Two_companies_iteration.xlsx')['Average_waiting_time [min]'].tolist()[0], pd.read_excel('Cooperation_scenario.xlsx')['Average_waiting_time [min]'].tolist()[0],pd.read_excel('Global_company.xlsx')['Average_waiting_time [min]'].tolist()[0]]
        plt.figure(figsize=(11, 6))
        plt.bar(labels, values, color=['red', 'blue', 'green'])
        plt.title('Average Waiting Time')
        plt.ylabel('Time in minutes')
        plt.savefig('Average_waiting_time.pdf')

        labels = [ 'Two_companies', 'Cooperation', 'One_company']
        values = [pd.read_excel('Two_companies_iteration.xlsx')['Total_pollution [T]'].tolist()[0], pd.read_excel('Cooperation_scenario.xlsx')['Total_pollution [T]'].tolist()[0], pd.read_excel('Global_company.xlsx')['Total_pollution [T]'].tolist()[0]]
        plt.figure(figsize=(11, 6))
        plt.bar(labels, values, color=['red', 'blue', 'green'])
        plt.title('Total Pollution')
        plt.ylabel('Pollution in T')
        plt.savefig('Total_pollution.pdf')

        labels = [ 'Two_companies', 'Cooperation', 'One_company']
        values = [pd.read_excel('Two_companies_iteration.xlsx')['Average_price [$]'].tolist()[0], pd.read_excel('Cooperation_scenario.xlsx')['Average_price [$]'].tolist()[0], pd.read_excel('Global_company.xlsx')['Average_price [$]'].tolist()[0]]
        plt.figure(figsize=(11, 6))
        plt.bar(labels, values, color=['red', 'blue', 'green'])
        plt.title('Average Price')
        plt.ylabel('Price in USD')
        plt.savefig('Average_price.pdf')

        two_company_data = [
            pd.read_excel('Two_companies_iteration.xlsx')['Percentage_taking_train'].tolist()[0],
            pd.read_excel('Two_companies_iteration.xlsx')['Percentage_walking'].tolist()[0],
            pd.read_excel('Two_companies_iteration.xlsx')['Percentage_taking_car'].tolist()[0]
        ]
        coop_data = [
            pd.read_excel('Cooperation_scenario.xlsx')['Percentage_taking_train'].tolist()[0],
            pd.read_excel('Cooperation_scenario.xlsx')['Percentage_walking'].tolist()[0],
            pd.read_excel('Cooperation_scenario.xlsx')['Percentage_taking_car'].tolist()[0]
        ]
        one_company_data = [
            pd.read_excel('Global_company.xlsx')['Percentage_taking_train'].tolist()[0],
            pd.read_excel('Global_company.xlsx')['Percentage_walking'].tolist()[0],
            pd.read_excel('Global_company.xlsx')['Percentage_taking_car'].tolist()[0]
        ]
        labels = ['Percentage Taking Train', 'Percentage Walking', 'Percentage Taking Car']
        x = range(len(labels))
        fig, ax = plt.subplots(figsize=(14, 6))
        ax.bar(x, two_company_data, width=0.25, color='red', label='Two Companies')
        ax.bar([i + 0.25 for i in x], coop_data, width=0.25, color='blue', label='Cooperation')
        ax.bar([i + 0.5 for i in x], one_company_data, width=0.25, color='green', label='One Company')
        ax.set_xticks([i + 0.2 for i in x])
        ax.set_xticklabels(labels)
        ax.set_ylabel('Percentage')
        ax.set_title('Percentage of Different Modes of Transportation')
        ax.legend()
        plt.savefig('percentage_transportation.pdf')

        plt.show()


    def calculate_global_train_lines(self):
        # This function will calculate the best train lines for the global company
        mapping = {node: node-1 for node in range(1, 25)}     
        all_train_lines = self.calculate_optimal_trajectories(range(1,25), mapping)
        two_companies_scenario = False
        n_improved = 20
        n_random = 600
        n_freq = 10
        n_trials = 90
        best_train_line, best_frequency, best_time_train_array, best_train_info_array, best_benefits = self.calculate_the_best_train_lines(3, all_train_lines, [10, 15], range(1,25), mapping, self.values_trips, n_improved, n_freq, n_trials, n_random, False, False, 0 , 0)
        is_it_valid, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, max_people_in_train, pourcentage_walking, pourcentage_taking_car, best_train_line, money_first_company, money_second_company = self.calculation(np.concatenate((best_frequency, best_frequency)), best_train_line, best_time_train_array, best_train_info_array, range(1,25), {node: node-1 for node in range(1, 25)}, self.values_trips)
        benefits_train_company = self.get_benefits(np.concatenate((best_frequency, best_frequency)), best_train_line, range(1,25), {node: node-1 for node in range(1, 25)}, self.values_trips, two_companies_scenario, 0, 0)[0]
        profits = money_first_company + money_second_company
        costs = self.get_costs_train_companies(np.concatenate((best_frequency, best_frequency)), best_train_line)
        self.load_existing_data('Global_company.xlsx')
        self.global_to_excel('Global_company.xlsx', benefits_train_company, profits, costs, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, pourcentage_walking, pourcentage_taking_car, [n_improved, n_freq, n_trials, n_random], best_train_line, best_frequency)

        print()
        print("FOR 1 GLOBAL COMPANY:")
        print("Benefits train company:", benefits_train_company)
        print("Maximum people in a train:", max_people_in_train)
        print("Average people per train:", average_people_per_train)
        print("Average travel + waiting time: (in min)", average_waiting_time)
        print("Average price: (in USD)", average_price)
        print("Total pollution: (in T)", total_pollution/1000000)
        print("Pourcentage of people taking train:", pourcentage_taking_train*100)
        print("Pourcentage of people walking:", pourcentage_walking*100)
        print("Pourcentage of people taking car:", pourcentage_taking_car*100)
        self.plotting_map(best_train_line, best_frequency + best_frequency, 'Global_company')
    
    def calculate_local_train_lines(self):
        # This function will calculate the best train lines for the 2 companies scenario
        mapping = {node: node-1 for node in range(1, 25)}
        nodes_top_company = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        shared_nodes = [10, 11, 12, 16, 18]
        nodes_bottom_company = [13, 14, 15, 17, 19, 20, 21, 22, 23, 24]
        node_to_index_top = {node: i for i, node in enumerate(nodes_top_company + shared_nodes)}
        node_to_index_bottom = {node: i for i, node in enumerate(nodes_bottom_company + shared_nodes)}
        values_trips_top_company = np.zeros((len(nodes_top_company)+len(shared_nodes), len(nodes_top_company)+len(shared_nodes)))
        values_trips_bottom_company = np.zeros((len(nodes_bottom_company)+len(shared_nodes), len(nodes_bottom_company)+len(shared_nodes)))
        G = nx.DiGraph()
        all_train_lines = self.calculate_optimal_trajectories(range(1,25), mapping)
        train_lines_top = []
        train_lines_bottom = [] 
        for connection in all_train_lines:
            starting_node = connection[0][0]
            arriving_node = connection[1][0]
            if (starting_node in nodes_top_company or starting_node in shared_nodes) and \
            (arriving_node in nodes_top_company or arriving_node in shared_nodes):
                train_lines_top.append(connection)
            if (starting_node in nodes_bottom_company or starting_node in shared_nodes) and \
            (arriving_node in nodes_bottom_company or arriving_node in shared_nodes):
                train_lines_bottom.append(connection)

        for item in all_train_lines:
            for source, target, time, passengers in zip(*item):
                G.add_edge(source, target, time=float(time), passengers=float(passengers))

        for i in range(len(self.values_trips)):
            i1 = i+1
            for j in range(len(self.values_trips)):
                j1 = j+1
                if (i1 in nodes_top_company and j1 in nodes_top_company) or (i1 in shared_nodes and j1 in nodes_top_company) or (i1 in nodes_top_company and j1 in shared_nodes):
                    values_trips_top_company[node_to_index_top[i1]][node_to_index_top[j1]] += self.values_trips[i][j]
                elif (i1 in nodes_bottom_company and j1 in nodes_bottom_company) or (i1 in shared_nodes and j1 in nodes_bottom_company) or (i1 in nodes_bottom_company and j1 in shared_nodes):
                    values_trips_bottom_company[node_to_index_bottom[i1]][node_to_index_bottom[j1]] += self.values_trips[i][j]
                elif i1 in shared_nodes and j1 in shared_nodes:
                    values_trips_top_company[node_to_index_top[i1]][node_to_index_top[j1]] = self.values_trips[i][j]/2
                    values_trips_bottom_company[node_to_index_bottom[i1]][node_to_index_bottom[j1]] += self.values_trips[i][j]/2
                elif i1 in nodes_top_company and j1 in nodes_bottom_company:
                    optimal_path, total_time, total_passengers = self.dijkstra_shortest_path(G, i1, j1)
                    first_shared_node, last_shared_node = self.find_transition_nodes(optimal_path, nodes_top_company, shared_nodes, nodes_bottom_company)
                    values_trips_top_company[node_to_index_top[i1]][node_to_index_top[first_shared_node]] +=  self.reduc_factor * self.values_trips[i][j]
                    if first_shared_node != last_shared_node:
                        values_trips_top_company[node_to_index_top[first_shared_node]][node_to_index_top[last_shared_node]] += self.reduc_factor * self.values_trips[i][j]/2
                        values_trips_bottom_company[node_to_index_bottom[first_shared_node]][node_to_index_bottom[last_shared_node]] += self.reduc_factor * self.values_trips[i][j]/2
                    values_trips_bottom_company[node_to_index_bottom[last_shared_node]][node_to_index_bottom[j1]] += self.reduc_factor * self.values_trips[i][j]
                elif i1 in nodes_bottom_company and j1 in nodes_top_company:
                    optimal_path, total_time, total_passengers = self.dijkstra_shortest_path(G, i1, j1)
                    first_shared_node, last_shared_node = self.find_transition_nodes(optimal_path, nodes_top_company, shared_nodes, nodes_bottom_company)
                    values_trips_bottom_company[node_to_index_bottom[i1]][node_to_index_bottom[first_shared_node]] += self.reduc_factor * self.values_trips[i][j]
                    if first_shared_node != last_shared_node:
                        values_trips_bottom_company[node_to_index_bottom[first_shared_node]][node_to_index_bottom[last_shared_node]] += self.reduc_factor * self.values_trips[i][j]/2
                        values_trips_top_company[node_to_index_top[first_shared_node]][node_to_index_top[last_shared_node]] += self.reduc_factor * self.values_trips[i][j]/2
                    values_trips_top_company[node_to_index_top[last_shared_node]][node_to_index_top[j1]] += self.reduc_factor * self.values_trips[i][j]


        n_improved = 20
        n_random =300
        n_freq = 10
        n_trials = 60  
        previous_bottom_best_train_line, previous_bottom_best_frequency, bottom_best_time_train_array, bottom_best_train_info_array, previous_bottom_benefits = self.calculate_the_best_train_lines(2, train_lines_bottom, [6, 11], nodes_bottom_company+shared_nodes, node_to_index_bottom, values_trips_bottom_company, n_improved, n_freq, n_trials, n_random, False, False, 0, 0)
        previous_top_best_train_line, previous_top_best_frequency, top_best_time_train_array, top_best_train_info_array, previous_top_benefits = self.calculate_the_best_train_lines(2, train_lines_top, [6, 11], nodes_top_company+shared_nodes, node_to_index_top, values_trips_top_company, n_improved, n_freq, n_trials, n_random, False, False, 0, 0)
        convergence= False
        ite = 0
        max_ite = 10
        while not convergence:  
            lines_bottom = range(len(previous_bottom_best_train_line))
            lines_top = range(len(previous_bottom_best_train_line), len(previous_bottom_best_train_line)+4)
            top_best_train_line, top_best_frequency, top_best_time_train_array, top_best_train_info_array, top_benefits = two_companies.calculate_the_best_local_train_lines([2,3], train_lines_top, [6, 11], nodes_top_company+shared_nodes, node_to_index_top, values_trips_top_company, n_improved, n_freq, n_trials, n_random, self.values_trips, previous_bottom_best_train_line, previous_bottom_best_frequency, 1, previous_top_best_train_line, lines_top, lines_bottom)
            lines_top = range(len(top_best_train_line))
            lines_bottom = range(len(top_best_train_line), len(top_best_train_line)+4)
            bottom_best_train_line, bottom_best_frequency, bottom_best_time_train_array, bottom_best_train_info_array, bottom_benefits = two_companies.calculate_the_best_local_train_lines([2,3], train_lines_bottom, [6, 11], nodes_bottom_company+shared_nodes, node_to_index_bottom, values_trips_bottom_company, n_improved, n_freq, n_trials, n_random, self.values_trips, top_best_train_line, top_best_frequency, 2, previous_bottom_best_train_line, lines_top, lines_bottom)
            ite+=1
            if ite == max_ite or (previous_bottom_benefits == bottom_benefits and previous_top_benefits == top_benefits):
                convergence = True
                print("End of while")
            previous_bottom_best_train_line = copy.deepcopy(bottom_best_train_line)
            previous_top_best_train_line = copy.deepcopy(top_best_train_line)
            previous_bottom_benefits = copy.deepcopy(bottom_benefits)
            previous_top_benefits = copy.deepcopy(top_benefits)
            previous_bottom_best_frequency = copy.deepcopy(bottom_best_frequency)
            print("end of ite:", ite)
        print()
        print("Final iteration:", ite)
        print()
     
        best_train_lines = []
        best_frequency = []
        best_frequency = top_best_frequency + bottom_best_frequency
        best_train_lines = top_best_train_line + bottom_best_train_line
        best_frequency = np.concatenate((best_frequency, best_frequency))
        mapping = {node: node-1 for node in range(1, 25)}
        time_train_array, train_info_array = self.get_array_time_train(best_frequency, best_train_lines, range(1,25), mapping, True, lines_top, lines_bottom)
        is_it_valid, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, max_people_in_train, pourcentage_walking, pourcentage_taking_car, best_train_lines, money_first_company, money_second_company = self.calculation(best_frequency, best_train_lines, time_train_array, train_info_array, range(1,25), mapping, self.values_trips)
        top_costs = self.get_costs_train_companies(top_best_frequency + top_best_frequency, top_best_train_line)
        bottom_costs = self.get_costs_train_companies(bottom_best_frequency + bottom_best_frequency, bottom_best_train_line)
        benefits_train_company = self.get_benefits(top_best_frequency + top_best_frequency + bottom_best_frequency + bottom_best_frequency, best_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, self.values_trips, True, lines_top, lines_bottom)

        self.load_existing_data('Two_companies_iteration.xlsx')
        self.iteration_to_excel('Two_companies_iteration.xlsx', average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, pourcentage_walking, pourcentage_taking_car, benefits_train_company[0], top_best_frequency, bottom_best_frequency, top_best_train_line, bottom_best_train_line, ite, benefits_train_company[1], top_costs + self.final_cost_top, benefits_train_company[2], bottom_costs + self.final_cost_bottom)
        self.plotting_map(best_train_lines, top_best_frequency + top_best_frequency + bottom_best_frequency + bottom_best_frequency, 'Two_companies_iteration')

        print()
        print("FOR 2 COMPANIES:")
        print()
        print("Benefits train company:", benefits_train_company)
        print("Maximum people in a train:", max_people_in_train)
        print("Average people per train:", average_people_per_train)
        print("Average travel + waiting time: (in min)", average_waiting_time)
        print("Average price: (in USD)", average_price)
        print("Total pollution: (in T)", total_pollution/1000000)
        print("Pourcentage of people taking train:", pourcentage_taking_train*100)
        print("Pourcentage of people walking:", pourcentage_walking*100)
        print("Pourcentage of people taking car:", pourcentage_taking_car*100)

    def change_values(self):
        # This function will change the values of the trips (increasing the proportion of inter-region trips)
        nodes_top_company = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        shared_nodes = [10, 11, 12, 16, 18]
        nodes_bottom_company = [13, 14, 15, 17, 19, 20, 21, 22, 23, 24]
        previous_number_of_people_inter_region = 0
        new_number_of_people_inter_region = 0
        previous_total_number_of_people = 0
        new_total_number_of_people = 0
        factor = 1
        for i in range(len(self.values_trips)):
            for j in range(len(self.values_trips)):
                previous_total_number_of_people += self.values_trips[i][j]
                new_total_number_of_people += self.values_trips[i][j]
                if (i in nodes_top_company and j in nodes_bottom_company) or (i in nodes_bottom_company and j in nodes_top_company):
                    previous_number_of_people_inter_region += self.values_trips[i][j]
                    new_total_number_of_people -= self.values_trips[i][j]
                    self.values_trips[i][j] = factor*self.values_trips[i][j]
                    new_number_of_people_inter_region += self.values_trips[i][j]
                    new_total_number_of_people += self.values_trips[i][j]
        for i in range(len(self.values_trips)):
            for j in range(len(self.values_trips)):
                self.values_trips[i][j] = self.values_trips[i][j]*previous_total_number_of_people/new_total_number_of_people
        people_inter_region = 0
        for i in range(len(self.values_trips)):
            for j in range(len(self.values_trips)):
                if (i in nodes_top_company and j in nodes_bottom_company) or (i in nodes_bottom_company and j in nodes_top_company):
                    people_inter_region += self.values_trips[i][j]

        previous_alpha = previous_number_of_people_inter_region/previous_total_number_of_people
        new_alpha = people_inter_region/previous_total_number_of_people
        print("Previous percentage of inter-region trips:", previous_alpha)
        print("New percentage:", new_alpha)

    
    def coop_scenario(self):
        previous_top_train_lines = ast.literal_eval(pd.read_excel('Two_companies_iteration.xlsx')['Top Company Lines'].tolist()[0])
        previous_top_frequency = ast.literal_eval(pd.read_excel('Two_companies_iteration.xlsx')['Top Company frequency'].tolist()[0])
        previous_bottom_train_lines = ast.literal_eval(pd.read_excel('Two_companies_iteration.xlsx')['Bottom Company Lines'].tolist()[0])
        previous_bottom_frequency = ast.literal_eval(pd.read_excel('Two_companies_iteration.xlsx')['Bottom Company frequency'].tolist()[0])

        n_trials = 60
        n_improved = 20
        n_random =300
        n_freq = 10
        convergence= False
        ite = 0
        max_ite = 10
        nodes =range(1, 25)
        mapping = {node: node-1 for node in range(1, 25)}
        all_train_lines = self.calculate_optimal_trajectories(range(1,25), mapping)
        previous_bottom_benefits = 0
        previous_top_benefits = 0
        while not convergence:  
            lines_bottom = range(len(previous_bottom_train_lines))
            lines_top = range(len(previous_bottom_train_lines), len(previous_bottom_train_lines)+4)            
            top_best_train_line, top_best_frequency, top_best_time_train_array, top_best_train_info_array, top_benefits = cooperation.iterate_train_lines([2,4], [6, 13], all_train_lines, nodes, mapping, n_improved, n_freq, n_trials, n_random, self.values_trips, previous_bottom_train_lines, previous_bottom_frequency, 1, previous_top_train_lines, lines_top, lines_bottom)
            lines_top = range(len(top_best_train_line))
            lines_bottom = range(len(top_best_train_line), len(top_best_train_line)+4)
            bottom_best_train_line, bottom_best_frequency, bottom_best_time_train_array, bottom_best_train_info_array, bottom_benefits = cooperation.iterate_train_lines([2,4], [6, 13], all_train_lines, nodes, mapping, n_improved, n_freq, n_trials, n_random, self.values_trips, top_best_train_line, top_best_frequency, 2, previous_bottom_train_lines, lines_top, lines_bottom)
            ite+=1
            if ite == max_ite or (previous_bottom_benefits == bottom_benefits and previous_top_benefits == top_benefits):
                convergence = True
                print("End of while")
            previous_bottom_best_train_line = copy.deepcopy(bottom_best_train_line)
            previous_top_best_train_line = copy.deepcopy(top_best_train_line)
            previous_bottom_benefits = copy.deepcopy(bottom_benefits)
            previous_top_benefits = copy.deepcopy(top_benefits)
            previous_bottom_frequency = copy.deepcopy(bottom_best_frequency)
            print("end of ite:", ite)
        print()
        print("Final iteration:", ite)
        print()
        best_train_lines = []
        best_frequency = []
        best_frequency = top_best_frequency + top_best_frequency + bottom_best_frequency + bottom_best_frequency
        best_train_lines = top_best_train_line + bottom_best_train_line
        mapping = {node: node-1 for node in range(1, 25)}
        time_train_array, train_info_array = self.get_array_time_train(best_frequency, best_train_lines, range(1,25), mapping, True, lines_top, lines_bottom)
        is_it_valid, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, max_people_in_train, pourcentage_walking, pourcentage_taking_car, best_train_lines, money_first_company, money_second_company = self.calculation(best_frequency, best_train_lines, time_train_array, train_info_array, range(1,25), mapping, self.values_trips)    
        top_costs = cooperation.coop_costs_top(top_best_frequency + top_best_frequency, top_best_train_line)[0]
        bottom_costs = cooperation.coop_costs_bottom(bottom_best_frequency + bottom_best_frequency, bottom_best_train_line)[0]
        benefits_train_company = self.get_benefits(best_frequency, best_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, self.values_trips, True, lines_top, lines_bottom)
        total_costs = self.get_costs_train_companies(best_frequency, best_train_lines)
        print("In Coop scenario")
        print("Benefits train company:", benefits_train_company[0])
        print("Top benefits:", top_benefits)
        print("Bottom benefits:", bottom_benefits)
        print(" Top costs:", top_costs)
        print(" Bottom costs:", bottom_costs)
        print("money first company:", money_first_company)
        print("money second company:", money_second_company)
        print("Total costs:", total_costs)
        print("Total costs coop calculus:", top_costs + bottom_costs)
        self.load_existing_data('Cooperation_scenario.xlsx')
        self.iteration_to_excel('Cooperation_scenario.xlsx', average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, pourcentage_walking, pourcentage_taking_car, benefits_train_company[0], top_best_frequency, bottom_best_frequency, top_best_train_line, bottom_best_train_line, ite, money_first_company, top_costs + self.final_cost_top, money_second_company, bottom_costs + self.final_cost_bottom)
        self.plotting_map(best_train_lines, best_frequency, 'Cooperation_scenario')

        print()
        print("FOR THE COOPERATION SCENARIO:")
        print()
        print("Benefits train company:", benefits_train_company)
        print("Maximum people in a train:", max_people_in_train)
        print("Average people per train:", average_people_per_train)
        print("Average travel + waiting time: (in min)", average_waiting_time)
        print("Average price: (in USD)", average_price)
        print("Total pollution: (in T)", total_pollution/1000000)
        print("Pourcentage of people taking train:", pourcentage_taking_train*100)
        print("Pourcentage of people walking:", pourcentage_walking*100)
        print("Pourcentage of people taking car:", pourcentage_taking_car*100)
    
    def solutionner(self):
        # Debugging function
        top_company_line = ast.literal_eval(pd.read_excel('Two_companies_iteration.xlsx')['Top Company Lines'].tolist()[0])
        top_company_frequency = ast.literal_eval(pd.read_excel('Two_companies_iteration.xlsx')['Top Company frequency'].tolist()[0])
        bottom_company_line = ast.literal_eval(pd.read_excel('Two_companies_iteration.xlsx')['Bottom Company Lines'].tolist()[0])
        bottom_company_frequency = ast.literal_eval(pd.read_excel('Two_companies_iteration.xlsx')['Bottom Company frequency'].tolist()[0])

        coop_top_company_line = ast.literal_eval(pd.read_excel('Cooperation_scenario.xlsx')['Top Company Lines'].tolist()[0])
        coop_top_company_frequency = ast.literal_eval(pd.read_excel('Cooperation_scenario.xlsx')['Top Company frequency'].tolist()[0])
        coop_bottom_company_line = ast.literal_eval(pd.read_excel('Cooperation_scenario.xlsx')['Bottom Company Lines'].tolist()[0])
        coop_bottom_company_frequency = ast.literal_eval(pd.read_excel('Cooperation_scenario.xlsx')['Bottom Company frequency'].tolist()[0])
        global_company_line = ast.literal_eval(pd.read_excel('Global_company.xlsx')['Train lines'].tolist()[0])
        global_company_frequency = ast.literal_eval(pd.read_excel('Global_company.xlsx')['Frequency'].tolist()[0])
        ite_final_normal = pd.read_excel('Two_companies_iteration.xlsx')['Final iteration'].tolist()[0]
        ite_final_coop = pd.read_excel('Cooperation_scenario.xlsx')['Final iteration'].tolist()[0]
        mapping = {node: node-1 for node in range(1, 25)}
        all_train_lines = self.calculate_optimal_trajectories(range(1,25), mapping)

        ite_freq = top_company_frequency + top_company_frequency + bottom_company_frequency + bottom_company_frequency
        ite_lines = top_company_line + bottom_company_line
        time_train_array, train_info_array = self.get_array_time_train(ite_freq , ite_lines, range(1,25), mapping, True, range(len(top_company_line)), range(len(top_company_line), len(top_company_line)+4))

        is_it_valid, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, max_people_in_train, pourcentage_walking, pourcentage_taking_car, best_train_lines, money_first_company, money_second_company = self.calculation(ite_freq, ite_lines, time_train_array, train_info_array, range(1,25), mapping, self.values_trips)
        benf = self.get_benefits(ite_freq, ite_lines, range(1,25), mapping, self.values_trips, True, range(len(top_company_line)), range(len(top_company_line), len(top_company_line)+4))
        costs_top = cooperation.coop_costs_top(top_company_frequency + top_company_frequency, top_company_line)[0]
        costs_bottom = cooperation.coop_costs_bottom(bottom_company_frequency + bottom_company_frequency, bottom_company_line)[0]
        
        print("money first company:", money_first_company)
        print("money second company:", money_second_company)
        print("Benefits train company:", benf)
        print("Costs top company:", costs_top + self.final_cost_top)
        print("Costs bottom company:", costs_bottom + self.final_cost_bottom)
        print("Total costs:", costs_top + self.final_cost_top + costs_bottom + self.final_cost_bottom)
        print("Final money first company:", money_first_company - costs_top - self.final_cost_top)
        print("Final money second company:", money_second_company - costs_bottom - self.final_cost_bottom)
        print("Final benefits train company:", money_first_company + money_second_company - costs_top - self.final_cost_top - costs_bottom - self.final_cost_bottom)
        print("people taking train", pourcentage_taking_train)

        coop_freq = coop_top_company_frequency + coop_top_company_frequency + coop_bottom_company_frequency + coop_bottom_company_frequency
        coop_lines = coop_top_company_line + coop_bottom_company_line
        time_train_array, train_info_array = self.get_array_time_train(coop_freq, coop_lines, range(1,25), mapping, True, range(len(coop_top_company_line)), range(len(coop_top_company_line), len(coop_top_company_line)+4))
        is_it_valid, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, max_people_in_train, pourcentage_walking, pourcentage_taking_car, best_train_lines, money_first_company, money_second_company = self.calculation(coop_freq, coop_lines, time_train_array, train_info_array, range(1,25), mapping, self.values_trips)
        benefits_train_company = self.get_benefits(coop_freq, coop_lines, range(1,25), mapping, self.values_trips, True, range(len(coop_top_company_line)), range(len(coop_top_company_line), len(coop_top_company_line)+4))
        costs_top = cooperation.coop_costs_top(coop_top_company_frequency + coop_top_company_frequency, coop_top_company_line)
        costs_bottom = cooperation.coop_costs_bottom(coop_bottom_company_frequency + coop_bottom_company_frequency, coop_bottom_company_line)
        print()
        print("money first company:", money_first_company + costs_bottom[1])
        print("money second company:", money_second_company + costs_top[1])
        print("Benefits train company:", benefits_train_company)
        print("Costs top company:", costs_top[0] + self.final_cost_top)
        print("Costs bottom company:", costs_bottom[0] + self.final_cost_bottom)
        print("Total costs:", costs_top[0] + self.final_cost_top + costs_bottom[0] + self.final_cost_bottom)
        print("Final money first company:", money_first_company - costs_top[0] - self.final_cost_top + costs_bottom[1])
        print("Final money second company:", money_second_company - costs_bottom[0] - self.final_cost_bottom + costs_top[1])
        print("Final benefits train company:",money_first_company + money_second_company - costs_top[0] - self.final_cost_top - costs_bottom[0] - self.final_cost_bottom + costs_top[1] + costs_bottom[1])
        print("people taking train", pourcentage_taking_train)
        self.clear_excel('Cooperation_scenario.xlsx')
        self.load_existing_data('Cooperation_scenario.xlsx')
        self.iteration_to_excel('Cooperation_scenario.xlsx', average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, pourcentage_walking, pourcentage_taking_car, benefits_train_company[0], coop_top_company_frequency, coop_bottom_company_frequency, coop_top_company_line, coop_bottom_company_line, ite_final_coop, benefits_train_company[1] + costs_bottom[1], costs_top[0] + self.final_cost_top, benefits_train_company[2] + costs_top[1], costs_bottom[0] + self.final_cost_bottom)
        time_train_array, train_info_array = self.get_array_time_train(global_company_frequency + global_company_frequency, global_company_line, range(1,25), mapping, True, range(len(global_company_line)), range(len(global_company_line), len(global_company_line)+4))
        is_it_valid, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, average_people_per_train, max_people_in_train, pourcentage_walking, pourcentage_taking_car, best_train_lines, money_first_company, money_second_company = self.calculation(global_company_frequency + global_company_frequency, global_company_line, time_train_array, train_info_array, range(1,25), mapping, self.values_trips)
        benefits_train_company = self.get_benefits(global_company_frequency + global_company_frequency, global_company_line, range(1,25), mapping, self.values_trips, True, range(len(global_company_line)), range(len(global_company_line), len(global_company_line)+4))
        costs_top = self.get_costs_train_companies(global_company_frequency + global_company_frequency, global_company_line)
        print()
  
        print("Benefits train company:", money_first_company + money_second_company)
        print("Total costs:", costs_top + self.final_cost_top + self.final_cost_bottom)
        print("Final benef:", money_first_company + money_second_company- costs_top - self.final_cost_top - self.final_cost_bottom)
        print("people taking train", pourcentage_taking_train)

    
    def showing_plot(self):
        # This function will the plot of the 2 company scenario
        station_coordinates = {
            1: (0, 7), 2: (6, 7), 3: (0, 6), 4: (2, 6), 5: (4, 6), 6: (6, 6),
            7: (8, 5), 8: (6, 5), 9: (4, 5), 10: (4, 4), 11: (2, 4), 12: (0, 4),
            13: (0, 0), 14: (2, 2), 15: (4, 2), 16: (6, 4), 17: (6, 3), 18: (8, 4),
            19: (6, 2), 20: (6, 0), 21: (4, 0), 22: (4, 1), 23: (2, 1), 24: (2, 0)
        }

        plt.title('Network')
        nodes_top = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        shared_nodes = [10, 11, 12, 16, 18]
        bottom_company_nodes = [13, 14, 15, 17, 19, 20, 21, 22, 23, 24]
        for station in nodes_top:
            x, y = station_coordinates[station]
            plt.plot(x, y, 'ro', markersize=0.1)
            plt.text(x, y, str(station), color='red', ha='center', va='center')
        for station in shared_nodes:
            x, y = station_coordinates[station]
            plt.plot(x, y, 'go', markersize=0.1)
            plt.text(x, y, str(station), color='green', ha='center', va='center')
        for station in bottom_company_nodes:
            x, y = station_coordinates[station]
            plt.plot(x, y, 'bo', markersize=0.1)
            plt.text(x, y, str(station), color='blue', ha='center', va='center')

        plt.xlim(-1, 9)
        plt.ylim(-1, 8)
        plt.xlabel('X')
        plt.ylabel('Y')
        plt.gca().set_aspect('equal', adjustable='box')
        plt.show()

    def final_plots(self):
        # This function will plot the final results 
        x1 = [0.25, 0.5, 0.75, 1]
        y1 = [172.5, 169.1, 155.7, 162.9]
        top_line_y1 =182.7
        bottom_line_y1 = 204.8
        x2 = [0.25, 0.5, 0.75, 1]
        y2 = [176.7, 186.4, 184.7, 173.9]
        top_line_y2 = 187.5
        bottom_line_y2 = 190.5
        x3 = [0.25, 0.5, 0.75, 1]
        y3 = [201.7, 193.7, 167.1, 173.75]
        top_line_y3 = 191.9
        bottom_line_y3 = 224
        plt.figure(figsize=(10, 6))
        plt.plot(x1, y1, color='blue', marker= 'o', linestyle = '-', label='25% Inter-region trips')
        plt.axhline(y=top_line_y1, color='blue', linestyle='--')
        plt.axhline(y=bottom_line_y1, color='blue', linestyle=':')
        plt.plot(x2, y2, color='green', marker= 'o', linestyle = '-', label='50% Inter-region trips')
        plt.axhline(y=top_line_y2, color='green', linestyle='--')
        plt.axhline(y=bottom_line_y2, color='green', linestyle=':')
        plt.plot(x3, y3, color='red', marker= 'o', linestyle = '-', label='75% Inter-region trips')
        plt.axhline(y=top_line_y3, color='red', linestyle='--')
        plt.axhline(y=bottom_line_y3, color='red', linestyle=':')
        plt.xlabel('Proportion of revenue retained')
        plt.ylabel('Pollution generated (T)')
        plt.title('Pollution generated in function of alpha')
        main_legend = plt.legend(loc='lower left')
        legend_lines = [
            Line2D([0], [0], color='black', linestyle='--', label='Global Company Scenario'),
            Line2D([0], [0], color='black', linestyle=':', label='Non-Cooperating Scenario')
        ]

        plt.gca().add_artist(main_legend)
        plt.legend(handles=legend_lines, loc='lower right')
        plt.show()

    def main(self):
        self.calculate_construction_costs()
        self.change_values()
        self.calculate_global_train_lines()
        self.calculate_local_train_lines()  
        self.coop_scenario()      
        self.plotting_graph()

class Cooperation_scenario:
    # This class will be used to calculate the cooperation scenario
    def __init__(self):
        self.alpha = 0.5
        self.cost_per_connection = 0
        pass

    def iterate_train_lines(self, max_min_number_train_lines, max_min_station_per_line, all_train_lines, nodes, mapping, number_improved_train_lines, number_frequency_train_lines, number_trials, number_random_train_lines, all_trips, other_company_lines, other_company_frequency, top_or_bottom, previous_lines, lines_top, lines_bottom):
        # This function will iterate the train lines
        random_train_lines = [self.coop_generate_random_train_lines(2, all_train_lines, max_min_station_per_line, nodes, other_company_lines) for _ in range(number_random_train_lines)]
        frequency_trainlines = [5] * (2 * 2)
        random_train_lines.extend([previous_lines])
        if top_or_bottom == 1:
            lines_top = range(4,8)
            lines_bottom = range(4)
        if top_or_bottom == 2:
            lines_top = range(4)
            lines_bottom = range(4,8)
        time_train_arrays = []
        train_info_arrays = []

        for train_lines in random_train_lines:
            time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
            time_train_arrays.append(time_train_array)
            train_info_arrays.append(train_info_array)
        previous_time_train_array, previous_train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + previous_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
        time_train_arrays.append(previous_time_train_array)
        train_info_arrays.append(previous_train_info_array)

        benefits = [self.get_benefits_coop(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom, top_or_bottom)[top_or_bottom-1] for train_lines, time_train_array, train_info_array in zip(random_train_lines, time_train_arrays, train_info_arrays)]
        if top_or_bottom == 1:
            costs = [self.coop_costs_top(frequency_trainlines, train_lines)[0] for train_lines in random_train_lines]

        if top_or_bottom == 2:
            costs = [self.coop_costs_bottom(frequency_trainlines, train_lines)[0] for train_lines in random_train_lines]
        benefits = [benefit - cost for benefit, cost in zip(benefits, costs)] 
        print("Calculating among the nodes: ", nodes)
        print()
        print("Current maximum benefits:", max(benefits))
        print()
        top_indices = sorted(range(len(benefits)), key=lambda i: benefits[i], reverse=True)[:number_improved_train_lines]
        best_random_train_lines = [random_train_lines[i] for i in top_indices]
        best_time_train_arrays = [time_train_arrays[i] for i in top_indices]
        best_train_info_arrays = [train_info_arrays[i] for i in top_indices]
        print("Improving the best train lines...")
        print()
        improved_train_lines = []
        improved_time_train_arrays = []
        improved_train_info_arrays = []

        for train_lines, time_train_array, train_info_array in zip(best_random_train_lines, best_time_train_arrays, best_train_info_arrays):
            improved_train_line, time_train_array, train_info_array = cooperation.improve_coop_train_lines(frequency_trainlines, train_lines, other_company_frequency, other_company_lines, all_train_lines, mapping, nodes, 0, all_trips, top_or_bottom, True, lines_top, lines_bottom)
            improved_train_lines.append(improved_train_line)
            improved_time_train_arrays.append(time_train_array)
            improved_train_info_arrays.append(train_info_array)
        benefits = [self.get_benefits_coop(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom, top_or_bottom)[top_or_bottom-1] for train_lines, time_train_array, train_info_array in zip(improved_train_lines, improved_time_train_arrays, improved_train_info_arrays)]
        if top_or_bottom == 1:
            costs = [self.coop_costs_top(frequency_trainlines, train_lines)[0] for train_lines in improved_train_lines]
        if top_or_bottom == 2:
            costs = [self.coop_costs_bottom(frequency_trainlines, train_lines)[0] for train_lines in improved_train_lines]
        benefits = [benefit - cost for benefit, cost in zip(benefits, costs)] 
        top_indices = sorted(range(len(improved_train_lines)), key=lambda i: benefits[i], reverse=True)[:number_frequency_train_lines]
        best_improved_train_lines = [improved_train_lines[i] for i in top_indices]
        best_time_train_arrays = [improved_time_train_arrays[i] for i in top_indices]
        best_train_info_arrays = [improved_train_info_arrays[i] for i in top_indices]
        
        print("Benefits after improving:", max(benefits))
        print("Optimizing the frequencies...")
        print()
        optimal_frequencies = []
        best_benefits = []
        for train_lines, time_train_array, train_info_array in zip(best_improved_train_lines, best_time_train_arrays, best_train_info_arrays):
            frequencies, benefits = self.calculate_coop_optimal_frequency(frequency_trainlines, train_lines, other_company_frequency, other_company_lines, all_train_lines, mapping, nodes, 0, all_trips, top_or_bottom, number_trials, True, lines_top, lines_bottom)
            optimal_frequencies.append(frequencies)
            best_benefits.append(benefits)
        print("Maximum benefits train company:", max(best_benefits))
        print()

        best_index = best_benefits.index(max(best_benefits))
        best_train_line = best_improved_train_lines[best_index]
        best_time_train_array = best_time_train_arrays[best_index]
        best_train_info_array = best_train_info_arrays[best_index]
        best_frequency = optimal_frequencies[best_index]

        best_time_train_array, best_train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, best_frequency, best_frequency)), other_company_lines + best_train_line, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
    
        return best_train_line, best_frequency, best_time_train_array, best_train_info_array, max(best_benefits)

    def get_benefits_coop(self, frequency_trainlines, train_lines, stations, mapping, values_trips, two_companies_scenario, lines_top, lines_bottom, top_or_bottom):
        # This function will calculate the benefits of the train companies
        time_train_array, train_info_array  = Global.get_array_time_train(frequency_trainlines, train_lines, stations, mapping, two_companies_scenario, lines_top, lines_bottom)
        _,_,_,_, pourcentage_taking_train, _,_,_, money_first_company, money_second_company = self.calculation_coop(frequency_trainlines, train_lines, time_train_array, train_info_array, stations, mapping, values_trips, lines_top, lines_bottom)
        if top_or_bottom == 1:
            number_initial_lines = len(lines_bottom)
            train_lines_bottom = train_lines[:number_initial_lines]
            train_lines_top = train_lines[number_initial_lines:]
            frequency_bottom = frequency_trainlines[:number_initial_lines]
            frequency_top = frequency_trainlines[number_initial_lines:]
        if top_or_bottom == 2:
            number_initial_lines = len(lines_top)
            train_lines_top = train_lines[:number_initial_lines]
            train_lines_bottom = train_lines[number_initial_lines:]
            frequency_top = frequency_trainlines[:number_initial_lines]
            frequency_bottom = frequency_trainlines[number_initial_lines:]
        
        add_benef_bottom = self.coop_costs_top(frequency_top, train_lines_top)[1]

        add_benef_top = self.coop_costs_bottom(frequency_bottom, train_lines_bottom)[1]
        return money_first_company + add_benef_top, money_second_company + add_benef_bottom

    def coop_generate_random_train_lines(self, number_trainlines, all_train_lines, length, nodes, others_lines):
        # This function will generate the train lines randomly
        train_lines = []
        valid = True
        for i in range(number_trainlines):
            starting_station = random.choice(nodes)
            length_train_line = random.randint(length[0], length[1])
            init_node = starting_station
            train_line  =[]
            for j in range(length_train_line):
                sum_passenger = 0
                next_indexes = []
                number_passenger = []
                times = []
                for k, sublist in enumerate(all_train_lines):
                    if sublist[0][0] == init_node and sublist[1][0] not in [x[0][0] for x in train_line]: 
                        if not Global.is_node_in_train_lines(train_lines, sublist[1][0]):
                            next_indexes.append(sublist[1][0])
                            times.append(sublist[2][0])
                            number_passenger.append(sublist[3][0])
                if next_indexes == []:
                    for k, sublist in enumerate(all_train_lines):
                        if sublist[0][0] == init_node and sublist[1][0] not in [x[0][0] for x in train_line]:  
                            next_indexes.append(sublist[1][0])
                            times.append(sublist[2][0])
                            number_passenger.append(sublist[3][0])
                if next_indexes == []:
                    for k, sublist in enumerate(all_train_lines):
                        if sublist[0][0] == init_node:
                            next_indexes.append(sublist[1][0])
                            times.append(sublist[2][0])
                            number_passenger.append(sublist[3][0])
                    indexes_in_train_line = []
                    for idx, row in enumerate(train_line):
                        node = row[0][0]  
                        if node in next_indexes:
                            indexes_in_train_line.append(idx) 
                    min_index = min(indexes_in_train_line)
                    next_node = train_line[min_index][0][0]
                    next_node_index = next_indexes.index(next_node)
                    time = times[next_node_index]
                    row = [[init_node], [next_node], [time], [0]] 
                    train_line.append(row)
                    break
                else:
                    if len(next_indexes) > 1:
                        probabilities = number_passenger/sum(number_passenger)
                        next_node = np.random.choice(next_indexes, p = probabilities)
                        time_index = next_indexes.index(next_node)
                        time = times[time_index]
                    else:
                        next_node = next_indexes[0]
                        time = times[0]
                   
                    row = [[init_node], [next_node], [time], [0]] 
                    train_line.append(row)
                    init_node = next_node
            train_lines.append(train_line)
        train_lines = [[[subsublist[0][0] for subsublist in sublist],
                  [subsublist[1][0] for subsublist in sublist],
                  [subsublist[2][0] for subsublist in sublist]]
                 for sublist in train_lines]
        opposite_train_lines = []
        for train_line in train_lines:
            opposite_train_line = [train_line[1][::-1], train_line[0][::-1], train_line[2][::-1]]
            opposite_train_lines.append(opposite_train_line)
        train_lines.extend(opposite_train_lines)
        for sublist in train_lines:
            sublist.append([0] * len(sublist[0]))

        stations_in_train_lines = set()
        for line in train_lines:
            stations_in_train_lines.update(line[0])
            stations_in_train_lines.update(line[1])
        for line in others_lines:
            stations_in_train_lines.update(line[0])
            stations_in_train_lines.update(line[1])
        for i in nodes:
            if i not in stations_in_train_lines:
                valid = False
                break
        if not valid: 
            return self.coop_generate_random_train_lines(number_trainlines, all_train_lines, length, nodes, others_lines)                     
        return train_lines
    
    def calculate_coop_optimal_frequency(self, frequency_trainlines, train_lines, other_company_frequency, other_company_lines, all_train_lines, mapping, stations, local_trips, all_trips, top_or_bottom, number_trials, two_companies_scenario, lines_top, lines_bottom):
        optuna.logging.set_verbosity(optuna.logging.ERROR)
        def objective(trial):
            frequencies = []
            for i in range(len(train_lines)//2):
                if trial.number == 0:  # First trial
                    frequency = trial.suggest_int(f'frequency_{i}', 5, 5)
                else:
                    frequency = trial.suggest_int(f'frequency_{i}', 3, 10)
                frequencies.append(frequency)
            all_frequencies = np.concatenate((other_company_frequency, other_company_frequency, frequencies, frequencies))
            if top_or_bottom == 1:
                costs = self.coop_costs_top(np.concatenate((frequencies,frequencies)), train_lines)[0]
                lines_bottom = range(4)
                lines_top = range(4,8)
            if top_or_bottom == 2:
                costs = self.coop_costs_bottom(np.concatenate((frequencies,frequencies)), train_lines)[0]
                lines_top = range(4)
                lines_bottom = range(4,8)
            benefits = self.get_benefits_coop(all_frequencies, other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom, top_or_bottom )[top_or_bottom-1] - costs
            return benefits
        study = optuna.create_study(direction='maximize')
        study.optimize(objective, n_trials=number_trials)
        best_frequencies = [study.best_params[f'frequency_{i}'] for i in range(len(train_lines)//2)]
        best_benefits = study.best_value
        return best_frequencies, best_benefits


    def calculation_coop(self, frequency_trainlines, train_lines, time_train_array, train_info_array, stations, mapping, values_trips, lines_top, lines_bottom):
        # This function will calculate the average waiting time for the users
        number_station = len(stations)
        self.total_number_of_people = np.sum(values_trips)
        total_waiting_time = 0
        total_price = 0
        total_pollution = 0
        total_number_people_taking_train = 0
        total_people_walking = 0
        total_people_taking_car = 0
        total_number = 0
        is_it_valid = True
        money_top_company = 0
        money_bottom_company = 0 
        total_money = 0
        nodes_top_company = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        nodes_bottom_company = [13, 14, 15, 17, 19, 20, 21, 22, 23, 24]
        shared_nodes = [10, 11, 12, 16, 18]
        for sublist in train_lines:
            sublist[3] = [0] * len(sublist[3])
        for i in stations:
            i =mapping[i]
            for j in stations:
                j= mapping[j]
                if i != j:
                    euclidean = others_network.get_euclidean_distance(i, j)
                    number_company = np.count_nonzero(train_info_array[i*number_station+j][2] == 1)
                    pourcentage, walking, time_train, price_train, time_alternative, price_alternative = Global.calculate_expected_pourcentage_taking_train(i, j, time_train_array, mapping, number_company)
                    
                    counter_train_to_take = np.count_nonzero(train_info_array[i*number_station+j][0] != -1)
                    total_distance_top = 0
                    total_distance_bottom = 0
                    distance_top_company_in_bottom_company = 0
                    distance_bottom_company_in_top_company = 0

                    for k in range (counter_train_to_take):
                        indexes_in_first_line = [index for index, num in enumerate(train_lines[train_info_array[i*number_station+j][0][k]][0]) if num == train_info_array[i*number_station+j][1][k][0]]
                        indexes_in_second_line = [index for index, num in enumerate(train_lines[train_info_array[i*number_station+j][0][k]][1]) if num == train_info_array[i*number_station+j][1][k][1]]
                        if indexes_in_second_line:
                            max_index_in_second_line = max(indexes_in_second_line)
                        if indexes_in_first_line:
                            min_index_in_first_line = min(indexes_in_first_line)
                        for l in range(min_index_in_first_line, max_index_in_second_line+1):
                           
                            if train_info_array[i*number_station+j][0][k] in lines_top:
                                total_distance_top += others_network.get_euclidean_distance(mapping[train_lines[train_info_array[i*number_station+j][0][k]][0][l]], mapping[train_lines[train_info_array[i*number_station+j][0][k]][1][l]])
                                if train_lines[train_info_array[i*number_station+j][0][k]][0][l] in nodes_bottom_company or train_lines[train_info_array[i*number_station+j][0][k]][1][l] in nodes_bottom_company:
                                    distance_top_company_in_bottom_company += others_network.get_euclidean_distance(mapping[train_lines[train_info_array[i*number_station+j][0][k]][0][l]], mapping[train_lines[train_info_array[i*number_station+j][0][k]][1][l]])
                            if train_info_array[i*number_station+j][0][k] in lines_bottom: 
                                total_distance_bottom += others_network.get_euclidean_distance(mapping[train_lines[train_info_array[i*number_station+j][0][k]][0][l]], mapping[train_lines[train_info_array[i*number_station+j][0][k]][1][l]])
                                if train_lines[train_info_array[i*number_station+j][0][k]][0][l] in nodes_top_company or train_lines[train_info_array[i*number_station+j][0][k]][1][l] in nodes_top_company:
                                    distance_bottom_company_in_top_company += others_network.get_euclidean_distance(mapping[train_lines[train_info_array[i*number_station+j][0][k]][0][l]], mapping[train_lines[train_info_array[i*number_station+j][0][k]][1][l]])
                    if time_train == np.inf:
                        is_it_valid = False
                    
                    total_waiting_time += pourcentage*time_train*values_trips[i][j] + (1-pourcentage)*time_alternative*values_trips[i][j]
                    if number_company >2 or number_company == 0:
                        is_it_valid = False
                    total_price += pourcentage*price_train*values_trips[i][j] + (1-pourcentage)*price_alternative*values_trips[i][j]
                    if is_it_valid == True:
                        ratio_full_top = (total_distance_top - distance_top_company_in_bottom_company)/(total_distance_top + total_distance_bottom)
                        ratio_full_bottom = (total_distance_bottom - distance_bottom_company_in_top_company)/(total_distance_top + total_distance_bottom)
                        ratio_share_top = distance_top_company_in_bottom_company/(total_distance_top + total_distance_bottom)
                        ratio_share_bottom = distance_bottom_company_in_top_company/(total_distance_top + total_distance_bottom)            
                        money_top_company += pourcentage*price_train*values_trips[i][j]*ratio_full_top + pourcentage*price_train*values_trips[i][j]*ratio_share_top*self.alpha +pourcentage*price_train*values_trips[i][j]*ratio_share_bottom*(1-self.alpha)
                        money_bottom_company += pourcentage*price_train*values_trips[i][j]*ratio_full_bottom + pourcentage*price_train*values_trips[i][j]*ratio_share_bottom*self.alpha + pourcentage*price_train*values_trips[i][j]*ratio_share_top*(1-self.alpha)
                        total_money += pourcentage*price_train*values_trips[i][j]
                    total_pollution += pourcentage * Global.pollution_train * values_trips[i][j] * euclidean * Global.euclidean_to_km

                    total_number_people_taking_train += pourcentage * values_trips[i][j]
                    if walking: 
                        total_people_walking += (1-pourcentage) * values_trips[i][j]
                    if not walking:
                        total_pollution += (1 - pourcentage) * Global.pollution_car * values_trips[i][j] * euclidean * Global.euclidean_to_km   
                        total_people_taking_car += (1-pourcentage) * values_trips[i][j] 
        average_waiting_time = total_waiting_time/self.total_number_of_people
        average_price = total_price/self.total_number_of_people
        pourcentage_taking_train = total_number_people_taking_train/self.total_number_of_people
        pourcentage_walking = total_people_walking/self.total_number_of_people
        pourcentage_taking_car = total_people_taking_car/self.total_number_of_people

        max_people_in_train = float('-inf')  
        total_people = 0
        total_trains = 0
        return is_it_valid, average_waiting_time, average_price, total_pollution, pourcentage_taking_train, pourcentage_walking, pourcentage_taking_car, train_lines, money_top_company, money_bottom_company

    def coop_costs_top(self, frequency_trainlines, train_lines):
        # This function will calculate the costs of the top company
        k = -1
        nodes_top_company = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        nodes_bottom_company = [13, 14, 15, 17, 19, 20, 21, 22, 23, 24]
        shared_nodes = [10, 11, 12, 16, 18]
        cost_train_company = 0
        benefits_other_company = 0
        
        for train_line in train_lines:
            k+=1
            total_km_line = 0
            total_time_line = np.sum(train_line[2])
            n_hours_per_day = 7
            for i in range(len(train_line[0])):
                init_node = train_line[0][i]
                term_node = train_line[1][i]
                if init_node in nodes_bottom_company or term_node in nodes_bottom_company:
                    cost_train_company += self.cost_per_connection
                    benefits_other_company += self.cost_per_connection
                euclidean = others_network.get_euclidean_distance(init_node-1, term_node-1)
                total_km_line += euclidean * Global.euclidean_to_km
            ratio_km_per_min = total_km_line/total_time_line
            cost_train_company += ratio_km_per_min * Global.cost_train_per_km * n_hours_per_day * 60 * 60/frequency_trainlines[k]
        cost_train_company+= len(train_lines) * Global.cost_per_trainline
        return cost_train_company, benefits_other_company
    
    def coop_costs_bottom(self, frequency_trainlines, train_lines):
        # This function will calculate the costs of the bottom company
        k = -1
        nodes_top_company = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        nodes_bottom_company = [13, 14, 15, 17, 19, 20, 21, 22, 23, 24]
        shared_nodes = [10, 11, 12, 16, 18]
        cost_train_company = 0
        benefits_other_company = 0
        n_hours_per_day = 7
        for train_line in train_lines:
            k+=1
            total_km_line = 0
            total_time_line = np.sum(train_line[2])
            for i in range(len(train_line[0])):
                init_node = train_line[0][i]
                term_node = train_line[1][i]
                if init_node in nodes_top_company or term_node in nodes_top_company:
                    cost_train_company += self.cost_per_connection
                    benefits_other_company += self.cost_per_connection
                euclidean = others_network.get_euclidean_distance(init_node-1, term_node-1)
                total_km_line += euclidean * Global.euclidean_to_km
            ratio_km_per_min = total_km_line/total_time_line
            cost_train_company += ratio_km_per_min * Global.cost_train_per_km * n_hours_per_day * 60 * 60/frequency_trainlines[k]
        cost_train_company+= len(train_lines) * Global.cost_per_trainline
        return cost_train_company, benefits_other_company

    def improve_coop_train_lines(self, frequency_trainlines, train_lines, other_company_frequency, other_company_lines, all_train_lines, mapping, stations, local_trips, all_trips, top_or_bottom, two_companies_scenario, lines_top, lines_bottom):
        # This function will improve the train lines
        num_train_lines = len(train_lines)
        half_index = num_train_lines // 2
        time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
        benefits = self.get_benefits_coop(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom, top_or_bottom)[top_or_bottom-1] 
        if top_or_bottom == 1:
            costs = self.coop_costs_top(frequency_trainlines, train_lines)[0]
        if top_or_bottom == 2:
            costs = self.coop_costs_bottom(frequency_trainlines, train_lines)[0] 
        best_benefits = benefits - costs
        best_reversed_train_lines = copy.deepcopy(train_lines[half_index:])
        current_reversed_train_lines = copy.deepcopy(train_lines[half_index:])
        for i in range(half_index):
            improving = True
            while improving:
                improving = False
                init_node = current_reversed_train_lines[i][1][len(current_reversed_train_lines[i][0])-1]
                next_indexes = []
                times = []
                for k, sublist in enumerate(all_train_lines):
                    if sublist[0][0] == init_node and sublist[1][0] not in current_reversed_train_lines[i][0][1:] and sublist[1][0] not in current_reversed_train_lines[i][1][1:] :
                        next_indexes.append(sublist[1][0])
                        times.append(sublist[2][0])
                if next_indexes == []:
                    continue
                for j in range(len(next_indexes)):
                    row = [[init_node], [next_indexes[j]], [times[j]], [0]]
                    current_reversed_train_lines[i][0].append(init_node) 
                    current_reversed_train_lines[i][1].append(next_indexes[j])
                    current_reversed_train_lines[i][2].append(times[j])
                    current_reversed_train_lines[i][3].append(0)
                    tested_train_lines = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(current_reversed_train_lines)))
                    time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
                    
                    benefits = self.get_benefits_coop(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom, top_or_bottom)[top_or_bottom-1] 
                    if top_or_bottom == 1:
                        costs = self.coop_costs_top(frequency_trainlines, tested_train_lines)[0]
                    if top_or_bottom == 2:
                        costs = self.coop_costs_bottom(frequency_trainlines, tested_train_lines)[0] 
                    current_benefits = benefits - costs
                    
                    if current_benefits > best_benefits:
                        improving = True
                        best_reversed_train_lines = copy.deepcopy(current_reversed_train_lines)
                        best_benefits = current_benefits
                        break
                    else:  
                        current_reversed_train_lines = copy.deepcopy(best_reversed_train_lines)
        network = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(best_reversed_train_lines)))
        best_first_half_train_lines = copy.deepcopy(network[half_index:])
        current_first_half_train_lines = copy.deepcopy(network[:half_index:])
        for i in range(half_index):
            improving = True
            while improving:
                improving = False
                init_node = current_first_half_train_lines[i][1][len(current_first_half_train_lines[i][0])-1]
                next_indexes = []
                times = []
                for k, sublist in enumerate(all_train_lines):
                    if sublist[0][0] == init_node and sublist[1][0] not in current_first_half_train_lines[i][0][1:] and sublist[1][0] not in current_first_half_train_lines[i][1][1:] :
                        next_indexes.append(sublist[1][0])
                        times.append(sublist[2][0])
                if next_indexes == []:
                    continue
                for j in range(len(next_indexes)):
                    row = [[init_node], [next_indexes[j]], [times[j]], [0]]
                    current_first_half_train_lines[i][0].append(init_node) 
                    current_first_half_train_lines[i][1].append(next_indexes[j])
                    current_first_half_train_lines[i][2].append(times[j])
                    current_first_half_train_lines[i][3].append(0)
                    tested_train_lines = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines)))
                    time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
                    benefits = self.get_benefits_coop(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom, top_or_bottom)[top_or_bottom-1] 
                    if top_or_bottom == 1:
                        costs = self.coop_costs_top(frequency_trainlines, tested_train_lines)[0]
                    if top_or_bottom == 2:
                        costs = self.coop_costs_bottom(frequency_trainlines, tested_train_lines)[0] 
                    current_benefits = benefits - costs
                    if current_benefits > best_benefits: 
                        improving = True
                        best_first_half_train_lines = copy.deepcopy(current_first_half_train_lines)
                        best_benefits = current_benefits
                        break
                    else:  
                        current_first_half_train_lines = copy.deepcopy(best_first_half_train_lines)
            improving = True
            while improving:
                removed_first_connection = [sublist[1:] for sublist in best_first_half_train_lines[i]]
                current_first_half_train_lines[i] = removed_first_connection
                tested_train_lines = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines)))
                time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)                
                is_it_valid =  Global.calculation(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, time_train_array, train_info_array, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips)[0]
                if is_it_valid:
                    benefits = self.get_benefits_coop(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom, top_or_bottom)[top_or_bottom-1] 
                    if top_or_bottom == 1:
                        costs = self.coop_costs_top(frequency_trainlines, tested_train_lines)[0]
                    if top_or_bottom == 2:
                        costs = self.coop_costs_bottom(frequency_trainlines, tested_train_lines)[0] 
                    current_benefits = benefits - costs
                else:
                    current_benefits = -50000
                if current_benefits > best_benefits and is_it_valid:
                    best_first_half_train_lines = copy.deepcopy(current_first_half_train_lines)
                    best_benefits = current_benefits
                else:
                    current_first_half_train_lines = copy.deepcopy(best_first_half_train_lines)
                    improving = False
            improving = True
            while improving:
                removed_last_connection = [sublist[:-1] for sublist in best_first_half_train_lines[i]]
                current_first_half_train_lines[i] = removed_last_connection
                tested_train_lines = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines)))
                time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
                is_it_valid =  Global.calculation(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, time_train_array, train_info_array, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips)[0]
                if is_it_valid:
                    benefits = self.get_benefits_coop(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom, top_or_bottom)[top_or_bottom-1] 
                    if top_or_bottom == 1:
                        costs = self.coop_costs_top(frequency_trainlines, tested_train_lines)[0]
                    if top_or_bottom == 2:
                        costs = self.coop_costs_bottom(frequency_trainlines, tested_train_lines)[0] 
                    current_benefits = benefits - costs
                else:
                    current_benefits = -50000
                if current_benefits > best_benefits and is_it_valid:
                    best_first_half_train_lines = copy.deepcopy(current_first_half_train_lines)
                    best_benefits = current_benefits
                else:
                    current_first_half_train_lines = copy.deepcopy(best_first_half_train_lines)
                    improving = False
        time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(best_first_half_train_lines))), range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
        return copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(best_first_half_train_lines))), time_train_array, train_info_array
 

class Two_Companies:
    # This class will calculate the best train lines for the non-cooperating scenario
    def __init__(self):
        pass   

    def calculate_the_best_local_train_lines(self, max_min_number_train_lines, local_train_lines, max_min_station_per_line, nodes, mapping, local_trips, number_improved_train_lines, number_frequency_train_lines, number_trials, number_random_train_lines, all_trips, other_company_lines, other_company_frequency, top_or_bottom, previous_lines, lines_top, lines_bottom):
        # This function will calculate the best train lines for the non-cooperating scenario
        random_train_lines = [Global.generate_random_train_lines(2, local_train_lines, max_min_station_per_line, nodes) for _ in range(number_random_train_lines)]
        frequency_trainlines = [5] * (2 * 2)
        if top_or_bottom == 1:
            lines_top = range(4,8)
            lines_bottom = range(4)
        if top_or_bottom == 2:
            lines_top = range(4)
            lines_bottom = range(4,8)
        
        random_train_lines.extend([previous_lines])
        time_train_arrays = []
        train_info_arrays = []
        for train_lines in random_train_lines:
            time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
            time_train_arrays.append(time_train_array)
            train_info_arrays.append(train_info_array)
        previous_time_train_array, previous_train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + previous_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
        time_train_arrays.append(previous_time_train_array)
        train_info_arrays.append(previous_train_info_array)
        benefits = [Global.get_benefits(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom)[top_or_bottom] for train_lines, time_train_array, train_info_array in zip(random_train_lines, time_train_arrays, train_info_arrays)]
        costs = [Global.get_costs_train_companies(frequency_trainlines, train_lines) for train_lines in random_train_lines]
        benefits = [benefit - cost for benefit, cost in zip(benefits, costs)] 
        print()
        print("Calculating among the nodes: ", nodes)
        print()
        print("Current maximum benefits:", max(benefits))
        print()

        top_indices = sorted(range(len(benefits)), key=lambda i: benefits[i], reverse=True)[:number_improved_train_lines]
        best_random_train_lines = [random_train_lines[i] for i in top_indices]
        best_time_train_arrays = [time_train_arrays[i] for i in top_indices]
        best_train_info_arrays = [train_info_arrays[i] for i in top_indices]

        print("Improving the best train lines...")
        print()
        improved_train_lines = []
        improved_time_train_arrays = []
        improved_train_info_arrays = []
        for train_lines, time_train_array, train_info_array in zip(best_random_train_lines, best_time_train_arrays, best_train_info_arrays):
            improved_train_line, time_train_array, train_info_array = self.improve_local_train_lines(frequency_trainlines, train_lines, other_company_frequency, other_company_lines, local_train_lines, mapping, nodes, local_trips, all_trips, top_or_bottom, True, lines_top, lines_bottom)
            improved_train_lines.append(improved_train_line)
            improved_time_train_arrays.append(time_train_array)
            improved_train_info_arrays.append(train_info_array)
        
        benefits = [Global.get_benefits(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom)[top_or_bottom] for train_lines, time_train_array, train_info_array in zip(improved_train_lines, improved_time_train_arrays, improved_train_info_arrays)]
        costs = [Global.get_costs_train_companies(frequency_trainlines, train_lines) for train_lines in improved_train_lines]
        benefits = [benefit - cost for benefit, cost in zip(benefits, costs)] 
        print()
        top_indices = sorted(range(len(improved_train_lines)), key=lambda i: benefits[i], reverse=True)[:number_frequency_train_lines]
        best_improved_train_lines = [improved_train_lines[i] for i in top_indices]
        best_time_train_arrays = [improved_time_train_arrays[i] for i in top_indices]
        best_train_info_arrays = [improved_train_info_arrays[i] for i in top_indices]
        print("Benefits after improving:", max(benefits))
        print("Optimizing the frequencies...")
        print()
        optimal_frequencies = []
        best_benefits = []
        for train_lines, time_train_array, train_info_array in zip(best_improved_train_lines, best_time_train_arrays, best_train_info_arrays):

            frequencies, benefits = self.calculate_local_optimal_frequency(frequency_trainlines, train_lines, other_company_frequency, other_company_lines, local_train_lines, mapping, nodes, local_trips, all_trips, top_or_bottom, number_trials, True, lines_top, lines_bottom)
            optimal_frequencies.append(frequencies)
            best_benefits.append(benefits)
        print("Maximum benefits train company:", max(best_benefits))
        print()

        best_index = best_benefits.index(max(best_benefits))
        best_train_line = best_improved_train_lines[best_index]
        best_time_train_array = best_time_train_arrays[best_index]
        best_train_info_array = best_train_info_arrays[best_index]
        best_frequency = optimal_frequencies[best_index]
        best_time_train_array, best_train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, best_frequency, best_frequency)), other_company_lines + best_train_line, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
        return best_train_line, best_frequency, best_time_train_array, best_train_info_array, max(best_benefits)

    def improve_local_train_lines(self, frequency_trainlines, train_lines, other_company_frequency, other_company_lines, all_train_lines, mapping, stations, local_trips, all_trips, top_or_bottom, two_companies_scenario, lines_top, lines_bottom):
        num_train_lines = len(train_lines)
        half_index = num_train_lines // 2
        
        time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
        benefits = Global.get_benefits(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom)[top_or_bottom]
        cost = Global.get_costs_train_companies(frequency_trainlines, train_lines) 
        best_benefits = benefits-cost
        best_reversed_train_lines = copy.deepcopy(train_lines[half_index:])
        current_reversed_train_lines = copy.deepcopy(train_lines[half_index:])
        for i in range(half_index):
            improving = True
            while improving:
                improving = False
                init_node = current_reversed_train_lines[i][1][len(current_reversed_train_lines[i][0])-1]
                next_indexes = []
                times = []
                for k, sublist in enumerate(all_train_lines):
                    if sublist[0][0] == init_node and sublist[1][0] not in current_reversed_train_lines[i][0][1:] and sublist[1][0] not in current_reversed_train_lines[i][1][1:] :
                        next_indexes.append(sublist[1][0])
                        times.append(sublist[2][0])
                if next_indexes == []:
                    continue
                for j in range(len(next_indexes)):
                    row = [[init_node], [next_indexes[j]], [times[j]], [0]]
                    current_reversed_train_lines[i][0].append(init_node) 
                    current_reversed_train_lines[i][1].append(next_indexes[j])
                    current_reversed_train_lines[i][2].append(times[j])
                    current_reversed_train_lines[i][3].append(0)
                    tested_train_lines = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(current_reversed_train_lines)))
                    time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
                    current_benefits = Global.get_benefits(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom)[top_or_bottom] - Global.get_costs_train_companies(frequency_trainlines, tested_train_lines) 
                    if current_benefits > best_benefits:
                        improving = True
                        best_reversed_train_lines = copy.deepcopy(current_reversed_train_lines)
                        best_benefits = current_benefits
                        break
                    else:  
                        current_reversed_train_lines = copy.deepcopy(best_reversed_train_lines)
        network = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(best_reversed_train_lines)))
        best_first_half_train_lines = copy.deepcopy(network[half_index:])
        current_first_half_train_lines = copy.deepcopy(network[:half_index:])
        for i in range(half_index):
            improving = True
            while improving:
                improving = False
                init_node = current_first_half_train_lines[i][1][len(current_first_half_train_lines[i][0])-1]
                next_indexes = []
                times = []
                for k, sublist in enumerate(all_train_lines):
                    if sublist[0][0] == init_node and sublist[1][0] not in current_first_half_train_lines[i][0][1:] and sublist[1][0] not in current_first_half_train_lines[i][1][1:] :
                        next_indexes.append(sublist[1][0])
                        times.append(sublist[2][0])
                if next_indexes == []:
                    continue
                for j in range(len(next_indexes)):
                    row = [[init_node], [next_indexes[j]], [times[j]], [0]]
                    current_first_half_train_lines[i][0].append(init_node) 
                    current_first_half_train_lines[i][1].append(next_indexes[j])
                    current_first_half_train_lines[i][2].append(times[j])
                    current_first_half_train_lines[i][3].append(0)
                    tested_train_lines = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines)))
                    time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
                    current_benefits = Global.get_benefits(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom)[top_or_bottom] - Global.get_costs_train_companies(frequency_trainlines, tested_train_lines) 
                    if current_benefits > best_benefits: 
                        improving = True
                        best_first_half_train_lines = copy.deepcopy(current_first_half_train_lines)
                        best_benefits = current_benefits
                        break
                    else:  
                        current_first_half_train_lines = copy.deepcopy(best_first_half_train_lines)
            improving = True
            while improving:
                removed_first_connection = [sublist[1:] for sublist in best_first_half_train_lines[i]]
                current_first_half_train_lines[i] = removed_first_connection
                tested_train_lines = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines)))
                time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)                
                is_it_valid =  Global.calculation(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, time_train_array, train_info_array, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips)[0]
                if is_it_valid:
                    current_benefits = Global.get_benefits(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom)[top_or_bottom] - Global.get_costs_train_companies(frequency_trainlines, tested_train_lines) 
                else:
                    current_benefits = -50000
                if current_benefits > best_benefits and is_it_valid:
                    best_first_half_train_lines = copy.deepcopy(current_first_half_train_lines)
                    best_benefits = current_benefits
                else:
                    current_first_half_train_lines = copy.deepcopy(best_first_half_train_lines)
                    improving = False
            improving = True
            while improving:
                removed_last_connection = [sublist[:-1] for sublist in best_first_half_train_lines[i]]
                current_first_half_train_lines[i] = removed_last_connection
                tested_train_lines = copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(current_first_half_train_lines)))
                time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
                is_it_valid =  Global.calculation(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, time_train_array, train_info_array, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips)[0]
                if is_it_valid:
                    current_benefits = Global.get_benefits(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + tested_train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom)[top_or_bottom] - Global.get_costs_train_companies(frequency_trainlines, tested_train_lines) 
                else:
                    current_benefits = -50000
                if current_benefits > best_benefits and is_it_valid:
                    best_first_half_train_lines = copy.deepcopy(current_first_half_train_lines)
                    best_benefits = current_benefits
                else:
                    current_first_half_train_lines = copy.deepcopy(best_first_half_train_lines)
                    improving = False
        time_train_array, train_info_array = Global.get_array_time_train(np.concatenate((other_company_frequency, other_company_frequency, frequency_trainlines)), other_company_lines + copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(best_first_half_train_lines))), range(1,25), {node: node-1 for node in range(1, 25)}, True, lines_top, lines_bottom)
        return copy.deepcopy(Global.construct_reversed_train_lines(copy.deepcopy(best_first_half_train_lines))), time_train_array, train_info_array
    
    def calculate_local_optimal_frequency(self, frequency_trainlines, train_lines, other_company_frequency, other_company_lines, all_train_lines, mapping, stations, local_trips, all_trips, top_or_bottom, number_trials, two_companies_scenario, lines_top, lines_bottom):
        optuna.logging.set_verbosity(optuna.logging.ERROR)

        def objective(trial):
            frequencies = []
            for i in range(len(train_lines)//2):
                if trial.number == 0:  # First trial
                    frequency = trial.suggest_int(f'frequency_{i}', 5, 5)
                else:
                    frequency = trial.suggest_int(f'frequency_{i}', 3, 10)
                frequencies.append(frequency)
            
            all_frequencies = np.concatenate((other_company_frequency, other_company_frequency, frequencies, frequencies))
            if top_or_bottom == 1:
                lines_top = range(4,8)
                lines_bottom = range(4)
            if top_or_bottom ==2:
                lines_bottom = range(4,8)
                lines_top = range(4)
            benefits = Global.get_benefits(all_frequencies, other_company_lines + train_lines, range(1,25), {node: node-1 for node in range(1, 25)}, all_trips, True, lines_top, lines_bottom)[top_or_bottom] - Global.get_costs_train_companies(np.concatenate((frequencies,frequencies)),  train_lines) 
            costs = Global.get_costs_train_companies(np.concatenate((frequencies,frequencies)),  train_lines)
            return benefits
        study = optuna.create_study(direction='maximize')
        study.optimize(objective, n_trials=number_trials)
        best_frequencies = [study.best_params[f'frequency_{i}'] for i in range(len(train_lines)//2)]
        best_benefits = study.best_value
        return best_frequencies, best_benefits

        
others_network = Others_Network()
Global = Global()
two_companies = Two_Companies()
cooperation = Cooperation_scenario()
Global.main()

